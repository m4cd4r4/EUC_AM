import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl import Workbook
import tkinter.font as tkFont

# Attempt to load the workbook or create it if it doesn't exist
try:
    workbook = load_workbook('EUC_Build_Room.xlsx')
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Item", "LastCount", "NewCount"])  # Create headers if not present
    workbook.save('EUC_Build_Room.xlsx')

# Function to update the Treeview widget with the spreadsheet data
def update_treeview():
    for row in tree.get_children():
        tree.delete(row)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)

# Function to update counts in the spreadsheet
def update_count(operation):
    try:
        # Find the row for the selected item
        selected_item = combobox_assets.get()
        input_value = int(entry_value.get())
        item_found = False

        for row in sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value == selected_item:
                item_found = True
                # Update LastCount with the value from NewCount
                row[1].value = row[2].value
                
                # Perform the operation
                if operation == 'add':
                    row[2].value = (row[2].value or 0) + input_value
                elif operation == 'subtract':
                    row[2].value = (row[2].value or 0) - input_value
                
                # Save the workbook
                workbook.save('EUC_Build_Room.xlsx')
                print(f"Updated '{selected_item}' with new count {row[2].value}")
                update_treeview()  # Update the Treeview with new data
                break

        if not item_found:
            print(f"Item '{selected_item}' not found in the spreadsheet.")

    except Exception as e:
        print(f"An error occurred: {e}")

# Initialize Tkinter root
root = tk.Tk()
root.title("IT Store Room Asset Management")
root.geometry("800x600")

# Define a large font for buttons
button_font = tkFont.Font(size=16)

# Create a frame to hold the widgets
frame = tk.Frame(root)
frame.pack(padx=10, pady=10, fill='both', expand=True)

# Dropdown for assets
items = [
    "HP USB External DVDRW Drive", "Wired Keyboard", "HP Laptop 360",
    "HP Laptop 840 G9", "HP Laptop 840 G10", "HP Docks G4", "HP Laptop Charger",
    "HP Desktop Mini", "Wireless Keyboard and Mice", "Wired Poly Headset 3325",
    "Poly Wireless headset", "34 inch curved monitor", "24 inch monitor"
]
combobox_assets = ttk.Combobox(frame, values=items, width=50)
combobox_assets.pack(pady=20)

# Entry field and buttons layout
entry_frame = tk.Frame(frame)
entry_frame.pack(pady=10)

button_subtract = tk.Button(entry_frame, text="-", command=lambda: update_count('subtract'), font=button_font)
button_subtract.pack(side=tk.LEFT, padx=5)

entry_value = tk.Entry(entry_frame, width=10)
entry_value.pack(side=tk.LEFT)

button_add = tk.Button(entry_frame, text="+", command=lambda: update_count('add'), font=button_font)
button_add.pack(side=tk.LEFT, padx=5)

# Treeview for displaying spreadsheet data
columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings")
for col in columns:
    tree.heading(col, text=col)
tree.pack(expand=True, fill="both", padx=10, pady=10)

# Initially populate the Treeview
update_treeview()

# Start the GUI event loop
root.mainloop()
