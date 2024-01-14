import customtkinter as ctk
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
import tkinter as tk
from tkinter import simpledialog, ttk

# Initialize Tkinter root with CustomTkinter
root = ctk.CTk()
root.title("Perth EUC Assets ")
root.geometry("800x600")

# Load the workbook or create it if it doesn't exist
workbook_path = 'EUC_Perth_Assets.xlsx'
if os.path.exists(workbook_path):
    workbook = load_workbook(workbook_path)
else:
    workbook = Workbook()
    workbook.active.title = '4.2_Items'
    workbook.create_sheet('4.2_Timestamps')
    workbook.create_sheet('BR_Items')
    workbook.create_sheet('BR_Timestamps')
    workbook['4.2_Items'].append(["Item", "LastCount", "NewCount"])
    workbook['4.2_Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['BR_Items'].append(["Item", "LastCount", "NewCount"])
    workbook['BR_Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook.save(workbook_path)

# Sheets reference
sheets = {
    'original': ('4.2_Items', '4.2_Timestamps'),
    'backup': ('BR_Items', 'BR_Timestamps')
}

# Currently active sheets
current_sheets = sheets['original']

# Define a larger font for Treeview
style = ttk.Style()
larger_font = ('Helvetica', 12)  # Increase font size by 50%
style.configure("Treeview", font=larger_font)

# Custom dialog class for SAN input
class SANInputDialog(simpledialog.Dialog):
    def body(self, master):
        self.entry = ttk.Entry(master)
        self.entry.pack()
        return self.entry

    def apply(self):
        san_input = self.entry.get()
        self.result = "SAN" + san_input

# Function to show SAN input dialog
def show_san_input():
    dialog = SANInputDialog(root, "Enter SAN Number")
    return dialog.result

# Function to update the Treeview widget with the spreadsheet data
def update_treeview():
    tree.delete(*tree.get_children())  # Clear the existing treeview entries
    for row in workbook[current_sheets[0]].iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)

# Function to log the changes to the log sheet and update the log view
def log_change(item, action, target_sheet, san_number=""):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    target_sheet.append([timestamp, item, action, san_number])
    workbook.save(workbook_path)
    update_log_view()

# Function to update the log view with the 5 most recent changes in ascending order
def update_log_view():
    log_view.delete(*log_view.get_children())  # Clear existing entries
    log_sheet = workbook[current_sheets[1]]
    all_rows = list(log_sheet.iter_rows(min_row=2, max_row=log_sheet.max_row, values_only=True))
    sorted_rows = sorted(all_rows, key=lambda r: r[0] if r[0] else datetime.min)
    for row in sorted_rows[-5:]:
        log_view.insert('', 0, values=row)  # Inserting at index 0 to put the most recent at the top

# Function to switch between original and backup sheets
def switch_sheets(sheet_type):
    global current_sheets
    current_sheets = sheets[sheet_type]
    update_treeview()
    update_log_view()

def update_count(operation):
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if selected_item:
        try:
            input_value = int(entry_value.get())
            item_sheet = workbook[current_sheets[0]]
            log_sheet = workbook[current_sheets[1]]

            # Check if the item is a laptop or mini-pc
            if 'laptop' in selected_item.lower() or 'mini-pc' in selected_item.lower():
                for _ in range(input_value):
                    san_number = show_san_input()
                    log_change(selected_item, f"{operation.capitalize()} 1", log_sheet, san_number)

            # Find the row for the selected item
            for row in item_sheet.iter_rows(min_row=2):
                if row[0].value == selected_item:
                    # Update LastCount with the former NewCount
                    row[1].value = row[2].value or 0

                    # Update NewCount based on the operation
                    if operation == 'add':
                        row[2].value = row[1].value + input_value
                    elif operation == 'subtract':
                        row[2].value = row[1].value - input_value
                    break

            workbook.save(workbook_path)
            update_treeview()
        except ValueError as e:
            print(f"Invalid input for count update: {e}")


# Create a frame to hold the widgets using CustomTkinter
frame = ctk.CTkFrame(root)
frame.pack(padx=10, pady=10, fill='both', expand=True)

# Entry field and buttons layout using CustomTkinter
entry_frame = ctk.CTkFrame(frame)
entry_frame.pack(pady=10)

# Button width calculation
button_width = 50  # Reduced width by 66%

# Sheet switch buttons
button_1 = ctk.CTkButton(entry_frame, text="4.2", command=lambda: switch_sheets('original'), width=button_width, font=("Helvetica", 16))
button_1.pack(side='left', padx=5)

button_2 = ctk.CTkButton(entry_frame, text="BR", command=lambda: switch_sheets('backup'), width=button_width, font=("Helvetica", 16))
button_2.pack(side='left', padx=5)

# Subtract button
button_subtract = ctk.CTkButton(entry_frame, text="-", command=lambda: update_count('subtract'), width=button_width, font=("Helvetica", 16))
button_subtract.pack(side='left', padx=5)

# Entry for count update
entry_value = tk.Entry(entry_frame, font=("Helvetica", 16), justify='center', width=10)
entry_value.pack(side='left')

# Add button
button_add = ctk.CTkButton(entry_frame, text="+", command=lambda: update_count('add'), width=button_width, font=("Helvetica", 16))
button_add.pack(side='left', padx=5)

# Treeview for item display
columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', height=8, style="Treeview")
for col in columns:
    tree.heading(col, text=col, anchor=tk.W)
    tree.column(col, anchor=tk.W, width=200, stretch=False)
tree.pack(expand=True, fill="both", padx=10, pady=20)
tree.bind('<ButtonRelease-1>', lambda e: entry_value.focus())

# Log view for displaying the 5 most recent changes
log_view_frame = ctk.CTkFrame(root)
log_view_frame.pack(side=tk.BOTTOM, fill='x', padx=10, pady=10)

log_view_columns = ("Timestamp", "Item", "Action", "SAN Number")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", height=5, style="Treeview")
for col in log_view_columns:
    log_view.heading(col, text=col, anchor=tk.W)
    log_view.column(col, anchor=tk.W, width=150, stretch=False)
log_view.pack(side='bottom', fill='x')

# Start the GUI event loop
root.after(100, update_treeview)
root.mainloop()
