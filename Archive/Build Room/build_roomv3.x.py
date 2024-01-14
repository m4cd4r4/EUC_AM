import logging.config
import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox
import customtkinter as ctk
from openpyxl import load_workbook, Workbook

class ApplicationLogic:
    def __init__(self, workbook_path):
        self.workbook_path = workbook_path
        self.workbook = self.load_or_create_workbook()
        self.all_sans_sheet = self.workbook['All SANs']
        self.sheets = {'original': ('4.2 Items', '4.2 Timestamps'), 'backup': ('BR Items', 'BR Timestamps')}
        self.current_sheets = self.sheets['original']

    def load_or_create_workbook(self):
        if Path(self.workbook_path).exists():
            return load_workbook(self.workbook_path)
        else:
            workbook = Workbook()
            # Code to create sheets and default data
            workbook.save(self.workbook_path)
            return workbook

    def is_san_unique(self, san_number):
        search_string = "SAN" + san_number if not san_number.startswith("SAN") else san_number
        unique = all(search_string != row[0] for row in self.all_sans_sheet.iter_rows(min_row=2, values_only=True))
        return unique

    def log_change(self, item, action, san_number="", timestamp_sheet=None):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            if timestamp_sheet is not None:
                san_number = f"SAN{san_number}" if san_number and not san_number.startswith('SAN') else san_number
                timestamp_sheet.append([timestamp, item, action, san_number])
                self.workbook.save(self.workbook_path)
                logging.info(f"Logged change: Time: {timestamp}, Item: {item}, Action: {action}, SAN: {san_number}")
            else:
                logging.error("No timestamp sheet provided for logging.")
        except Exception as e:
            logging.error(f"Failed to log change: {e}")
            raise

    def switch_sheets(self, sheet_type):
        self.current_sheets = self.sheets[sheet_type]

    def update_count(self, operation, selected_item, input_value, san_required):
        item_sheet = self.workbook[self.current_sheets[0]]
        timestamp_sheet = self.workbook[self.current_sheets[1]]

        for row in item_sheet.iter_rows(min_row=2):
            if row[0].value == selected_item:
                row[1].value = row[2].value or 0
                if operation == 'add':
                    row[2].value = (row[2].value or 0) + input_value
                elif operation == 'subtract':
                    row[2].value = max((row[2].value or 0) - input_value, 0)

        if not san_required:
            self.log_change(selected_item, operation, "", timestamp_sheet)

        self.workbook.save(self.workbook_path)

class ApplicationUI:
    def __init__(self, root, logic):
        self.root = root
        self.logic = logic
        self.create_widgets()

    def create_widgets(self):
        self.frame = ctk.CTkFrame(self.root)
        self.frame.pack(padx=10, pady=10, fill='both', expand=True)
        self.entry_frame = ctk.CTkFrame(self.frame)
        self.entry_frame.pack(pady=10)
        # ... More widgets here ...

    def update_treeview(self):
        self.tree.delete(*self.tree.get_children())
        item_sheet = self.logic.workbook[self.logic.current_sheets[0]]
        row_count = 0
        for row in item_sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                self.tree.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
                row_count += 1
        self.tree.tag_configure('oddrow', background='#f0f0f0')
        self.tree.tag_configure('evenrow', background='white')

    def update_log_view(self):
        if hasattr(self, 'log_view'):
            self.log_view.delete(*self.log_view.get_children())
            log_sheet = self.logic.workbook[self.logic.current_sheets[1]]
            all_rows = list(log_sheet.iter_rows(min_row=2, values_only=True))
            sorted_rows = sorted(all_rows, key=lambda r: datetime.strptime(r[0], "%Y-%m-%d %H:%M:%S") if r[0] else datetime.min, reverse=True)
            for row in sorted_rows:
                if row[0] is not None:
                    self.log_view.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
                    row_count += 1
            self.log_view.tag_configure('oddrow', background='#f0f0f0')
            self.log_view.tag_configure('evenrow', background='white')

    def toggle_theme(self):
        current_mode = ctk.get_appearance_mode()
        new_mode = "dark" if current_mode == "light" else "light"
        ctk.set_appearance_mode(new_mode)

    def open_spreadsheet(self):
        try:
            if os.name == 'nt':
                os.startfile(self.logic.workbook_path)
            else:
                opener = "open" if sys.platform == "darwin" else "xdg-open"
                subprocess.run([opener, self.logic.workbook_path])
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open the spreadsheet: {e}")

def main():
    root = ctk.CTk()
    root.title("Perth EUC Assets")
    root.geometry("550x600")

    script_directory = Path(__file__).parent
    workbook_path = script_directory / 'EUC_Perth_Assets.xlsx'

    logic = ApplicationLogic(workbook_path)
    ui = ApplicationUI(root, logic)

    root.mainloop()

if __name__ == "__main__":
    main()
