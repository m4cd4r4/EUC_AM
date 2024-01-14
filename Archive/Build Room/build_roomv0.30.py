import customtkinter as ctk
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
import tkinter as tk
from tkinter import ttk, simpledialog

# Initialize Tkinter root with CustomTkinter
root = ctk.CTk()
root.title("Store-Room 4.2")
root.geometry("800x600")

# Load the workbook or create it if it doesn't exist
workbook_path = 'EUC_Perth_Assets.xlsx'
if os.path.exists(workbook_path):
    workbook = load_workbook(workbook_path)
else:
    workbook = Workbook()
    workbook.active.title = '4.2_Items'
    workbook.create_sheet('4.2_Timestamps')
    workbook.create_sheet('BR_Items')
    workbook.create_sheet('BR_Timestamps')
    workbook['4.2_Items'].append(["Item", "LastCount", "NewCount"])
    workbook['4.2_Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['BR_Items'].append(["Item", "LastCount", "NewCount"])
    workbook['BR_Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook.save(workbook_path)

sheet = workbook['4.2_Items']
log_sheet = workbook['4.2_Timestamps']
br_sheet = workbook['BR_Items']
br_log_sheet = workbook['BR_Timestamps']

# Define a larger font for Treeview
style = ttk.Style()
larger_font = ('Helvetica', 12)  # Increase font size by 50%
style.configure("Treeview", font=larger_font)

# Custom dialog class for SAN input
class SANInputDialog(simpledialog.Dialog):
    # ... [SANInputDialog body]

# Function to show SAN input dialog
def show_san_input():
    # ... [show_san_input body]

# Function to update the Treeview widget with the spreadsheet data
def update_treeview():
    # ... [update_treeview body]

# Function to log the changes to the log sheet and update the log view
def log_change(item, action, target_sheet, san_number=""):
    # ... [log_change body]

# Function to update the log view with the 5 most recent changes in ascending order
def update_log_view():
    # ... [update_log_view body]

# Function to update counts in the spreadsheet
def update_count(operation):
    # ... [update_count body]

# Create a frame to hold the widgets using CustomTkinter
frame = ctk.CTkFrame(root)
frame.pack(padx=10, pady=10, fill='both', expand=True)

# Entry field and buttons layout using CustomTkinter
entry_frame = ctk.CTkFrame(frame)
entry_frame.pack(pady=10)

# Button width calculation
button_width = 50  # Reduced width by 66%

# Additional buttons (You can change the command and text as needed)
button_1 = ctk.CTkButton(entry_frame, text="Btn1", width=button_width, font=("Helvetica", 16))
button_1.pack(side='left', padx=5)

button_2 = ctk.CTkButton(entry_frame, text="Btn2", width=button_width, font=("Helvetica", 16))
button_2.pack(side='left', padx=5)

button_subtract = ctk.CTkButton(entry_frame, text="-", command=lambda: update_count('subtract'), width=button_width, font=("Helvetica", 16))
button_subtract.pack(side='left', padx=5)

entry_value = tk.Entry(entry_frame, font=("Helvetica", 16), justify='center', width=10)
entry_value.pack(side='left')

button_add = ctk.CTkButton(entry_frame, text="+", command=lambda: update_count('add'), width=button_width, font=("Helvetica", 16))
button_add.pack(side='left', padx=5)

# Treeview for displaying spreadsheet data
columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', height=8, style="Treeview")
for col in columns:
    tree.heading(col, text=col, anchor=tk.W)
    tree.column(col, anchor=tk.W, width=200, stretch=False)
tree.pack(expand=True, fill="both", padx=10, pady=20)
tree.bind('<ButtonRelease-1>', lambda e: entry_value.focus())

# Log view for displaying the 5 most recent changes
log_view_frame = ctk.CTkFrame(root)
log_view_frame.pack(side=tk.BOTTOM, fill='x', padx=10, pady=10)

log_view_columns = ("Timestamp", "Item", "Action", "SAN Number")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", height=5, style="Treeview")
for col in log_view_columns:
    log_view.heading(col, text=col, anchor=tk.W)
    log_view.column(col, anchor=tk.W, width=150, stretch=False)
log_view.pack(side='bottom', fill='x')

# Start the GUI event loop
root.after(100, update_treeview)
root.mainloop()
