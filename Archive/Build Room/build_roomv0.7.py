import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl import Workbook
import tkinter.font as tkFont
from datetime import datetime

# Attempt to load the workbook or create it if it doesn't exist
try:
    workbook = load_workbook('EUC_Build_Room.xlsx')
    sheet = workbook.active
    try:
        log_sheet = workbook['Sheet2']  # Attempt to access the second sheet
    except KeyError:
        log_sheet = workbook.create_sheet('Sheet2')  # Create the second sheet if it doesn't exist
        log_sheet.append(["Timestamp", "Item", "Action"])  # Create headers if not present
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Item", "LastCount", "NewCount"])  # Create headers if not present
    log_sheet = workbook.create_sheet('Sheet2')  # Create the second sheet
    log_sheet.append(["Timestamp", "Item", "Action"])  # Create headers if not present
    workbook.save('EUC_Build_Room.xlsx')

# Function to update the Treeview widget with the spreadsheet data
def update_treeview():
    for row in tree.get_children():
        tree.delete(row)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)

# Function to log the changes to the second sheet and update the log view
def log_change(item, action):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_sheet.append([timestamp, item, action])
    workbook.save('EUC_Build_Room.xlsx')
    update_log_view()

# Function to update the log view with the 5 most recent changes
def update_log_view():
    for row in log_view.get_children():
        log_view.delete(row)
    for row in log_sheet.iter_rows(min_row=max(1, log_sheet.max_row - 4), max_row=log_sheet.max_row, values_only=True):
        log_view.insert('', 'end', values=row)

# Function to update counts in the spreadsheet
def update_count(operation):
    try:
        # Find the row for the selected item
        selected_item = combobox_assets.get()
        input_value = int(entry_value.get())
        item_found = False

        for row in sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value == selected_item:
                item_found = True
                # Update LastCount with the value from NewCount
                row[1].value = row[2].value
                
                # Perform the operation
                new_value = (row[2].value or 0) + input_value if operation == 'add' else (row[2].value or 0) - input_value
                row[2].value = new_value
                
                # Save the workbook
                workbook.save('EUC_Build_Room.xlsx')
                print(f"Updated '{selected_item}' with new count {new_value}")
                update_treeview()  # Update the Treeview with new data
                
                # Log the change
                action = 'Add' if operation == 'add' else 'Subtract'
                log_change(selected_item, f"{action} {input_value}")
                break

        if not item_found:
            print(f"Item '{selected_item}' not found in the spreadsheet.")

    except Exception as e:
        print(f"An error occurred: {e}")

# Initialize Tkinter root
root = tk.Tk()
root.title("IT Store Room Asset Management")
root.geometry("800x600")

# Define a large font for buttons
button_font = tkFont.Font(size=16)

# Create a frame to hold the widgets
frame = tk.Frame(root)
frame.pack(padx=10, pady=10, fill='both', expand=True)

# Dropdown for assets
items = [
    "HP USB External DVDRW Drive", "Wired Keyboard", "HP Laptop 360",
    "HP Laptop 840 G9", "HP Laptop 840 G10", "HP Docks G4", "HP Laptop Charger",
    "HP Desktop Mini", "Wireless Keyboard and Mice", "Wired Poly Headset 3325",
    "Poly Wireless headset", "34 inch curved monitor", "24 inch monitor"
]
combobox_assets = ttk.Combobox(frame, values=items, width=50)
combobox_assets.pack(pady=20)

# Entry field and buttons layout
entry_frame = tk.Frame(frame)
entry_frame.pack(pady=10)

button_subtract = tk.Button(entry_frame, text="-", command=lambda: update_count
