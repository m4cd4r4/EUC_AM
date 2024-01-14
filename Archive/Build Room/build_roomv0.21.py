import customtkinter as ctk  # Import CustomTkinter
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
import os
import tkinter as tk
from tkinter import ttk

# Initialize Tkinter root with CustomTkinter
root = ctk.CTk()
root.title("Store-Room 4.2")
root.geometry("800x600")

# Load the workbook or create it if it doesn't exist
workbook_path = 'EUC_Build_Room.xlsx'
workbook = load_workbook(workbook_path) if os.path.exists(workbook_path) else Workbook()
sheet = workbook['Sheet1'] if 'Sheet1' in workbook.sheetnames else workbook.active
log_sheet = workbook['Sheet2'] if 'Sheet2' in workbook.sheetnames else workbook.create_sheet('Sheet2')

# Function to update the Treeview widget with the spreadsheet data
def update_treeview():
    tree.delete(*tree.get_children())  # Clear the existing treeview entries
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)

# Function to log the changes to the second sheet and update the log view
def log_change(item, action):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_sheet.append([timestamp, item, action])
    workbook.save(workbook_path)
    update_log_view()

# Function to update the log view with the 5 most recent changes in ascending order
def update_log_view():
    log_view.delete(*log_view.get_children())  # Clear existing entries
    all_rows = list(log_sheet.iter_rows(min_row=2, max_row=log_sheet.max_row, values_only=True))
    sorted_rows = sorted(all_rows, key=lambda r: r[0] if r[0] else datetime.min)
    for row in sorted_rows[-5:]:
        log_view.insert('', 0, values=row)  # Inserting at index 0 to put the most recent at the top

# Function to update counts in the spreadsheet
def update_count(operation):
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if selected_item:
        try:
            input_value = int(entry_value.get())
            for row in sheet.iter_rows(min_row=2, values_only=False):
                if row[0].value == selected_item:
                    row[1].value = row[2].value  # Update LastCount with current NewCount
                    new_value = row[2].value or 0
                    row[2].value = new_value + input_value if operation == 'add' else new_value - input_value
                    workbook.save(workbook_path)
                    update_treeview()
                    log_change(selected_item, f"{operation.capitalize()} {input_value}")
                    break
        except ValueError as e:
            print(f"Invalid input for count update: {e}")

# Create a frame to hold the widgets using CustomTkinter
frame = ctk.CTkFrame(root)
frame.pack(padx=10, pady=10, fill='both', expand=True)

# Entry field and buttons layout using CustomTkinter
entry_frame = ctk.CTkFrame(frame)
entry_frame.pack(pady=10)

# Subtract Button using CustomTkinter
button_subtract = ctk.CTkButton(entry_frame, text="-", command=lambda: update_count('subtract'), font=("Helvetica", 16))
button_subtract.pack(side='left', padx=5)

# Entry Value using CustomTkinter
entry_value = ctk.CTkEntry(entry_frame, width=200, font=("Helvetica", 16))
entry_value.pack(side='left')

# Add Button using CustomTkinter
button_add = ctk.CTkButton(entry_frame, text="+", command=lambda: update_count('add'), font=("Helvetica", 16))
button_add.pack(side='left', padx=5)

# Treeview for displaying spreadsheet data
columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', height=8)
for col in columns:
    tree.heading(col, text=col, anchor=tk.W)
    tree.column(col, anchor=tk.W, width=200, stretch=False)
tree.pack(expand=True, fill="both", padx=10, pady=20)
tree.bind('<ButtonRelease-1>', lambda e: entry_value.focus())

# Log view for displaying the 5 most recent changes
log_view_frame = ctk.CTkFrame(root)
log_view_frame.pack(side=tk.BOTTOM, fill='x', padx=10, pady=10)

log_view_columns = ("Timestamp", "Item", "Action")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", height=5)
for col in log_view_columns:
    log_view.heading(col, text=col, anchor=tk.W)
    log_view.column(col, anchor=tk.W, width=150, stretch=False)
log_view.pack(side='bottom', fill='x')

# Start the GUI event loop
root.after(100, update_treeview)
root.mainloop()
