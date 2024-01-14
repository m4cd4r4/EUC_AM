import customtkinter as ctk
import os
import tkinter as tk
from tkinter import simpledialog, ttk
from openpyxl import load_workbook, Workbook
from datetime import datetime
import subprocess  # Import subprocess for non-Windows systems

# Initialize Tkinter root with CustomTkinter
root = ctk.CTk()
root.title("Perth EUC Assets")
root.geometry("600x600")

# Load the workbook or create it if it doesn't exist
workbook_path = 'EUC_Perth_Assets.xlsx'
if os.path.exists(workbook_path):
    workbook = load_workbook(workbook_path)
else:
    # ... [Workbook initialization code remains unchanged]

# ... [Rest of the initialization code remains unchanged]

# Custom dialog class for SAN input
# ... [SANInputDialog class remains unchanged]

# Function to show SAN input dialog
# ... [show_san_input function remains unchanged]

# Function to open the spreadsheet
# ... [open_spreadsheet function remains unchanged]

# Create a frame to hold the widgets using CustomTkinter
# ... [Frame creation code remains unchanged]

# Entry field and buttons layout using CustomTkinter
# ... [Entry field and buttons layout code remains unchanged]

# Function to update the Treeview widget with the spreadsheet data
# ... [update_treeview function remains unchanged]

# Function to log the changes to the log sheet and update the log view
# ... [log_change function remains unchanged]

# Function to update the "All SANs" sheet
def update_all_sans_sheet(item, san_number, action):
    all_sans_sheet = workbook['All SANs']
    if action == 'add':
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        all_sans_sheet.append([item, san_number, timestamp])
    elif action == 'subtract':
        for row in all_sans_sheet.iter_rows(min_row=2):
            if row[0].value == item and row[1].value == san_number:
                all_sans_sheet.delete_rows(row[0].row)
                break

# Function to switch between original and backup sheets
# ... [switch_sheets function remains unchanged]

# Function to update count and handle SAN items
def update_count(operation):
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if selected_item:
        try:
            input_value = int(entry_value.get())
            item_sheet = workbook[current_sheets[0]]
            log_sheet = workbook[current_sheets[1]]

            for row in item_sheet.iter_rows(min_row=2):
                if row[0].value == selected_item:
                    # Update LastCount with the current NewCount
                    row[1].value = row[2].value or 0 

                    # Update NewCount based on the operation
                    if operation == 'add':
                        row[2].value = (row[2].value or 0) + input_value
                    elif operation == 'subtract':
                        row[2].value = max((row[2].value or 0) - input_value, 0)

                    # Log change, with SAN number for specific items
                    if any(keyword in selected_item.lower() for keyword in ['840', 'x360', 'desktop mini']):
                        for _ in range(input_value):
                            san_number = show_san_input()
                            if san_number:
                                log_change(selected_item, f"{operation.capitalize()} 1", log_sheet, san_number)
                                update_all_sans_sheet(selected_item, san_number, operation)
                            else:
                                break
                    else:
                        # Log change for items without a SAN number
                        log_change(selected_item, f"{operation.capitalize()} {input_value}", log_sheet)

                    break  # Exit the loop once the item is found and updated

            workbook.save(workbook_path)
            update_treeview()
        except ValueError as e:
            tk.messagebox.showerror("Error", f"Invalid input for count update: {e}")

# Treeview for item display
# ... [Treeview setup code remains unchanged]

# Log view for displaying the changes with scrollbar
# ... [Log view setup code remains unchanged]

# Start the GUI event loop
root.after(100, update_treeview)
root.mainloop()
