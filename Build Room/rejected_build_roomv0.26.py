# The "SAN Number" column is added to both the "4.2_Timestamps" and "BR_Timestamps" sheets.
# An additional entry field input_san is included for users to enter the SAN number when logging changes.
# The log_change function is modified to include the SAN number in the logged data.
# The log view now displays the SAN number as part of the logged information.

import customtkinter as ctk
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
import tkinter as tk
from tkinter import ttk

# Initialize Tkinter root with CustomTkinter
root = ctk.CTk()
root.title("Store-Room 4.2")
root.geometry("800x600")

# Load the workbook or create it if it doesn't exist
workbook_path = 'EUC_Perth_Assets.xlsx'
workbook = load_workbook(workbook_path) if os.path.exists(workbook_path) else Workbook()

# Check for existing sheets or create them
sheet_names = workbook.sheetnames
sheet = workbook['4.2_Items'] if '4.2_Items' in sheet_names else workbook.create_sheet('4.2_Items')
log_sheet = workbook['4.2_Timestamps'] if '4.2_Timestamps' in sheet_names else workbook.create_sheet('4.2_Timestamps')
br_sheet = workbook['BR_Items'] if 'BR_Items' in sheet_names else workbook.create_sheet('BR_Items')
br_log_sheet = workbook['BR_Timestamps'] if 'BR_Timestamps' in sheet_names else workbook.create_sheet('BR_Timestamps')

# Initialize headers if the workbook was just created
if not os.path.exists(workbook_path):
    sheet.append(["Item", "LastCount", "NewCount"])
    log_sheet.append(["Timestamp", "Item", "Action", "SAN Number"])
    br_sheet.append(["Item", "LastCount", "NewCount"])
    br_log_sheet.append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook.save(workbook_path)

# Define a larger font for Treeview
style = ttk.Style()
larger_font = ('Helvetica', 12)  # Increase font size by 50%
style.configure("Treeview", font=larger_font)

# Function to validate Entry input
def validate_input(P):
    if P.isdigit() or P == "":
        return True
    return False

# Function to update the Treeview widget with the spreadsheet data
def update_treeview():
    tree.delete(*tree.get_children())  # Clear the existing treeview entries
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)

# Function to log the changes to the log sheet and update the log view
def log_change(item, action, san_number):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_sheet.append([timestamp, item, action, san_number])
    br_log_sheet.append([timestamp, item, action, san_number])  # Also log to BR_Timestamps
    workbook.save(workbook_path)
    update_log_view()

# Function to update the log view with the 5 most recent changes in ascending order
def update_log_view():
    log_view.delete(*log_view.get_children())  # Clear existing entries
    all_rows = list(log_sheet.iter_rows(min_row=2, max_row=log_sheet.max_row, values_only=True))
    sorted_rows = sorted(all_rows, key=lambda r: r[0] if r[0] else datetime.min)
    for row in sorted_rows[-5:]:
        log_view.insert('', 0, values=row)  # Inserting at index 0 to put the most recent at the top

# Function to update counts in the spreadsheet
def update_count(operation):
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if selected_item:
        try:
            input_value = int(entry_value.get())
            san_number = input_san.get()  # Retrieve SAN number from another entry field
            for row in sheet.iter_rows(min_row=2, values_only=False):
                if row[0].value == selected_item:
                    row[1].value = row[2].value  # Update LastCount with current NewCount
                    new_value = row[2].value or 0
                    row[2].value = new_value + input_value if operation == 'add' else new_value - input_value
                    workbook.save(workbook_path)
                    update_treeview()
                    log_change(selected_item, f"{operation.capitalize()} {input_value}", san_number)
                    break
        except ValueError as e:
            print(f"Invalid input for count update: {e}")

# Create a frame to hold the widgets using CustomTkinter
frame = ctk.CTkFrame(root)
frame.pack(padx=10, pady=10, fill='both', expand=True)

# Entry field and buttons layout using CustomTkinter
entry_frame = ctk.CTkFrame(frame)
entry_frame.pack(pady=10)

button_subtract = ctk.CTkButton(entry_frame, text="-", command=lambda: update_count('subtract'), font=("Helvetica", 16))
button_subtract.pack(side='left', padx=5)

vcmd = (root.register(validate_input), '%P')
entry_value = tk.Entry(entry_frame, font=("Helvetica", 16), justify='center', width=10, validate='key', validatecommand=vcmd)
entry_value.pack(side='left')

# Additional entry field for SAN Number
input_san = tk.Entry(entry_frame, font=("Helvetica", 16), justify='center', width=10)
input_san.pack(side='left', padx=10)

button_add = ctk.CTkButton(entry_frame, text="+", command=lambda: update_count('add'), font=("Helvetica", 16))
button_add.pack(side='left', padx=5)

columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', height=8, style="Treeview")
for col in columns:
    tree.heading(col, text=col, anchor=tk.W)
    tree.column(col, anchor=tk.W, width=200, stretch=False)
tree.pack(expand=True, fill="both", padx=10, pady=20)
tree.bind('<ButtonRelease-1>', lambda e: entry_value.focus())

# Log view for displaying the 5 most recent changes
log_view_frame = ctk.CTkFrame(root)
log_view_frame.pack(side=tk.BOTTOM, fill='x', padx=10, pady=10)

log_view_columns = ("Timestamp", "Item", "Action", "SAN Number")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", height=5, style="Treeview")
for col in log_view_columns:
    log_view.heading(col, text=col, anchor=tk.W)
    log_view.column(col, anchor=tk.W, width=150, stretch=False)
log_view.pack(side='bottom', fill='x')

# Start the GUI event loop
root.after(100, update_treeview)
root.mainloop()