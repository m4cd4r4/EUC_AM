import customtkinter as ctk
import os
import tkinter as tk
from tkinter import simpledialog, ttk
from openpyxl import load_workbook, Workbook
from datetime import datetime
import subprocess  # Import subprocess for non-Windows systems

# Initialize Tkinter root with CustomTkinter
root = ctk.CTk()
root.title("Perth EUC Assets")
root.geometry("600x600")

# Load the workbook or create it if it doesn't exist
workbook_path = 'EUC_Perth_Assets.xlsx'
if os.path.exists(workbook_path):
    workbook = load_workbook(workbook_path)
else:
    workbook = Workbook()
    workbook.active.title = '4.2_Items'
    workbook.create_sheet('4.2 Timestamps')
    workbook.create_sheet('BR Items')
    workbook.create_sheet('BR Timestamps')
    workbook.create_sheet('Project Designated Items')  # Sheet 5
    workbook.create_sheet('Project Designated Timestamps')  # Sheet 6
    workbook.create_sheet('All_SANs')  # Sheet 7
    # Initialize headers
    workbook['4.2 Items'].append(["Item", "LastCount", "NewCount"])
    workbook['4.2 Timestamps'].append(["Item", "Action", "SAN Number", "Time"])  # Corrected sheet name
    workbook['BR Items'].append(["Item", "LastCount", "NewCount"])
    workbook['BR Timestamps'].append(["Item", "Action", "SAN Number", "Time"])  # Adjusted order
    workbook['Project Designated Items'].append(["Item", "LastCount", "NewCount"])  # Modify as needed
    workbook['Project Designated Timestamps'].append(["Item", "Action", "SAN Number", "Time"])  # Adjusted order
    workbook['All SANs'].append(["Item", "SAN Number", "Time"])  # Adjusted order


# Sheets reference
sheets = {
    'original': ('4.2 Items', '4.2 Timestamps'),
    'backup': ('BR Items', 'BR Timestamps')
}

# Currently active sheets
current_sheets = sheets['original']

# Define a larger font for Treeview
style = ttk.Style()
larger_font = ('Helvetica', 12)  # Increase font size by 50%
style.configure("Treeview", font=larger_font)

# Custom dialog class for SAN input
class SANInputDialog(simpledialog.Dialog):
    def body(self, master):
        # Validation command
        vcmd = (master.register(self.on_validate), '%P')
        self.entry = ttk.Entry(master, validate="key", validatecommand=vcmd)
        self.entry.pack()
        return self.entry

    def on_validate(self, P):
        # Allow only numeric input
        return P.isdigit() or P == ""

    def apply(self):
        san_input = "SAN" + self.entry.get()
        # Check for length between 5 and 6 digits (excluding "SAN" prefix)
        if 5 <= len(san_input) - 3 <= 6:
            self.result = san_input
        else:
            tk.messagebox.showerror("Error", "Please enter a valid SAN with 5 to 6 digits.")
            self.result = None

# Function to show SAN input dialog
def show_san_input():
    while True:
        dialog = SANInputDialog(root, "Enter SAN Number")
        san_input = dialog.result
        if san_input and len(san_input) >= 8:
            return san_input


# Function to open the spreadsheet
def open_spreadsheet():
    try:
        if os.name == 'nt':  # Windows
            os.startfile(workbook_path)
        else:  # macOS, Linux, etc.
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, workbook_path])
    except Exception as e:
        tk.messagebox.showerror("Error", f"Failed to open the spreadsheet: {e}")

# Create a frame to hold the widgets using CustomTkinter
frame = ctk.CTkFrame(root)
frame.pack(padx=10, pady=10, fill='both', expand=True)

# Entry field and buttons layout using CustomTkinter
entry_frame = ctk.CTkFrame(frame)
entry_frame.pack(pady=10)

# Button width calculation
button_width = 50  # Reduced width by 66%

# Sheet switch buttons
button_1 = ctk.CTkButton(entry_frame, text="Basement 4.2", command=lambda: switch_sheets('original'), width=button_width, font=("Helvetica", 14))
button_1.pack(side='left', padx=(10, 5))  # Adjusted left padding for a tasteful close position


button_2 = ctk.CTkButton(entry_frame, text="Build Room", command=lambda: switch_sheets('backup'), width=button_width, font=("Helvetica", 14))
button_2.pack(side='left', padx=5)

# Subtract button
button_subtract = ctk.CTkButton(entry_frame, text="-", command=lambda: update_count('subtract'), width=button_width, font=("Helvetica", 14))
button_subtract.pack(side='left', padx=5)

# Entry for count update
entry_value = tk.Entry(entry_frame, font=("Helvetica", 14), justify='center', width=4)
entry_value.pack(side='left')

# Add button
button_add = ctk.CTkButton(entry_frame, text="+", command=lambda: update_count('add'), width=button_width, font=("Helvetica", 14, "bold"))
button_add.pack(side='left', padx=5)

# .xlsx button
xlsx_button = ctk.CTkButton(entry_frame, text=".xlsx", command=open_spreadsheet, width=button_width, font=("Helvetica", 16))
xlsx_button.pack(side='left', padx=5)


# Function to update the Treeview widget with the spreadsheet data
def update_treeview():
    tree.delete(*tree.get_children())  # Clear the existing treeview entries
    for row in workbook[current_sheets[0]].iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)

# Function to log the changes to the log sheet and update the log view
def log_change(item, action, target_sheet, san_number=""):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if target_sheet.title in ['4.2 TImestamps', 'BR Timestamps', 'Project Designated Timestamps']:
        # Append data in the new order for BR Timestamps and Project Designated Timestamps
        target_sheet.append([item, action, san_number, timestamp])
    elif target_sheet.title == 'All SANs':
        # Append data in the new order for All_SANs
        target_sheet.append([item, san_number, timestamp])
    else:
        # For other sheets, keep the original order
        target_sheet.append([item, action, san_number, timestamp])
    

    # Debugging print statement
    print(f"Logging change: {item}, {action}, {san_number}")

    # # Update the All SANs sheet if applicable
    # if san_number:
    #     update_all_sans_sheet(item, san_number, 'add' if 'add' in action.lower() else 'subtract')


    workbook.save(workbook_path)
    update_log_view()

def update_log_view():
    log_view.delete(*log_view.get_children())  # Clear existing entries
    log_sheet = workbook[current_sheets[1]]
    all_rows = list(log_sheet.iter_rows(min_row=2, max_row=log_sheet.max_row, values_only=True))

    def get_datetime(value):
        try:
            return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
        except (TypeError, ValueError):
            return datetime.min

    sorted_rows = sorted(all_rows, key=lambda r: get_datetime(r[0]), reverse=True)
    for row in sorted_rows:
        log_view.insert('', '0', values=row)


# Function to update the "All SANs" sheet
def update_all_sans_sheet(item, san_number, action):
    all_sans_sheet = workbook['All SANs']
    # Debugging print statement
    print(f"Updating All SANs sheet: {item}, {san_number}, {action}")
    if action == 'add':
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        all_sans_sheet.append([item, san_number, timestamp])
    elif action == 'subtract':
        for row in all_sans_sheet.iter_rows(min_row=2):
            if row[0].value == item and row[1].value == san_number:
                all_sans_sheet.delete_rows(row[0].row)
                break


# Function to switch between original and backup sheets
def switch_sheets(sheet_type):
    global current_sheets
    current_sheets = sheets[sheet_type]
    update_treeview()
    update_log_view()

def update_count(operation):
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if selected_item:
        try:
            input_value = int(entry_value.get())
            item_sheet = workbook[current_sheets[0]]
            log_sheet = workbook[current_sheets[1]]

            for row in item_sheet.iter_rows(min_row=2):
                if row[0].value == selected_item:
                    # Update LastCount with the current NewCount
                    row[1].value = row[2].value or 0 

                    # Update NewCount based on the operation
                    if operation == 'add':
                        row[2].value = (row[2].value or 0) + input_value
                    elif operation == 'subtract':
                        row[2].value = max((row[2].value or 0) - input_value, 0)

                    # Log change, with SAN number for specific items
                    if any(keyword in selected_item.lower() for keyword in ['840', 'x360', 'desktop mini']):
                        for _ in range(input_value):
                            san_number = show_san_input()
                            if san_number:
                                log_change(selected_item, f"{operation.capitalize()} 1", log_sheet, san_number)
                                # Ensure update_all_sans_sheet is called here to avoid duplication
                                update_all_sans_sheet(selected_item, san_number, operation)
                            else:
                                break
                    else:
                        # Log change for items without a SAN number
                        log_change(selected_item, f"{operation.capitalize()} {input_value}", log_sheet)

                    break  # Exit the loop once the item is found and updated

            workbook.save(workbook_path)
            update_treeview()
        except ValueError as e:
            tk.messagebox.showerror("Error", f"Invalid input for count update: {e}")


# Function to update count and handle SAN items
def update_count(operation):
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if selected_item:
        try:
            input_value = int(entry_value.get())
            item_sheet = workbook[current_sheets[0]]
            log_sheet = workbook[current_sheets[1]]

            for row in item_sheet.iter_rows(min_row=2):
                if row[0].value == selected_item:
                    # Update LastCount with the current NewCount
                    row[1].value = row[2].value or 0 

                    # Update NewCount based on the operation
                    if operation == 'add':
                        row[2].value = (row[2].value or 0) + input_value
                    elif operation == 'subtract':
                        row[2].value = max((row[2].value or 0) - input_value, 0)

                    # Log change, with SAN number for specific items
                    if any(keyword in selected_item.lower() for keyword in ['840', 'x360', 'desktop mini']):
                        for _ in range(input_value):
                            san_number = show_san_input()
                            if san_number:
                                log_change(selected_item, f"{operation.capitalize()} 1", log_sheet, san_number)
                                update_all_sans_sheet(selected_item, san_number, operation)
                            else:
                                break
                    else:
                        # Log change for items without a SAN number
                        log_change(selected_item, f"{operation.capitalize()} {input_value}", log_sheet)

                    break  # Exit the loop once the item is found and updated

            workbook.save(workbook_path)
            update_treeview()
        except ValueError as e:
            tk.messagebox.showerror("Error", f"Invalid input for count update: {e}")


# Treeview for item display
columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', style="Treeview")
for col in columns:
    tree.heading(col, text=col, anchor='w')
    tree.column(col, anchor='w', width=200, stretch=False)
tree.pack(expand=True, fill="both", padx=10, pady=20)

# Log view for displaying the changes with scrollbar
log_view_frame = ctk.CTkFrame(root)
log_view_frame.pack(side=tk.BOTTOM, fill='both', expand=True, padx=10, pady=10)

log_view_columns = ("Item", "Action", "SAN Number", )
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", style="Treeview", height=8)
for col in log_view_columns:
    log_view.heading(col, text=col, anchor='w')
    log_view.column(col, anchor='w', width=150, stretch=False)

# Scrollbar for the Log View
scrollbar_log = ttk.Scrollbar(log_view_frame, orient="vertical", command=log_view.yview)
scrollbar_log.pack(side='right', fill='y')
log_view.configure(yscrollcommand=scrollbar_log.set)

log_view.pack(expand=True, fill='both')

# Start the GUI event loop
root.after(100, update_treeview)
root.mainloop()