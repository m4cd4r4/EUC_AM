import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl import Workbook
import tkinter.font as tkFont
from datetime import datetime
import os

# Initialize Tkinter root
root = tk.Tk()
root.title("Store-Room 4.2")
root.geometry("800x600")

# Define a large font for buttons
button_font = tkFont.Font(size=16)

# Load the workbook or create it if it doesn't exist
workbook_path = 'EUC_Build_Room.xlsx'
if not os.path.exists(workbook_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet1'
    sheet.append(["Item", "LastCount", "NewCount"])  # Create headers if not present
    log_sheet = workbook.create_sheet('Sheet2')  # Create the log sheet
    log_sheet.append(["Timestamp", "Item", "Action"])  # Create headers if not present
    workbook.save(workbook_path)
else:
    workbook = load_workbook(workbook_path)
    sheet = workbook['Sheet1'] if 'Sheet1' in workbook.sheetnames else workbook.active
    log_sheet = workbook['Sheet2'] if 'Sheet2' in workbook.sheetnames else workbook.create_sheet('Sheet2')

# Function to update the Treeview widget with the spreadsheet data
def update_treeview():
    tree.delete(*tree.get_children())  # Clear the existing treeview entries
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)

# Function to log the changes to the second sheet and update the log view
def log_change(item, action):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_sheet.append([timestamp, item, action])
    workbook.save(workbook_path)
    update_log_view()

# Function to update the log view with the 5 most recent changes
def update_log_view():
    for row in log_view.get_children():
        log_view.delete(row)
    for row in log_sheet.iter_rows(min_row=max(1, log_sheet.max_row - 4), max_row=log_sheet.max_row, values_only=True):
        log_view.insert('', 'end', values=row)

# Function to update counts in the spreadsheet
def update_count(operation):
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if selected_item:
        input_value = int(entry_value.get())
        for row in sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value == selected_item:
                current_count = row[2].value or 0
                row[2].value = current_count + input_value if operation == 'add' else current_count - input_value
                workbook.save(workbook_path)
                update_treeview()
                log_change(selected_item, f"{operation.capitalize()} {input_value}")
                break

# Create a frame to hold the widgets
frame = tk.Frame(root)
frame.pack(padx=10, pady=10, fill='both', expand=True)

# Entry field and buttons layout
entry_frame = tk.Frame(frame)
entry_frame.pack(pady=10)

# Subtract Button
button_subtract = tk.Button(entry_frame, text="-", command=lambda: update_count('subtract'), font=button_font)
button_subtract.pack(side=tk.LEFT, padx=5)

# Entry Value
entry_value = tk.Entry(entry_frame, width=10, font=button_font)
entry_value.pack(side=tk.LEFT)

# Add Button
button_add = tk.Button(entry_frame, text="+", command=lambda: update_count('add'), font=button_font)
button_add.pack(side=tk.LEFT, padx=5)

# Treeview for displaying spreadsheet data
columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', height=8)
for col in columns:
    tree.heading(col, text=col, anchor=tk.W)
    tree.column(col, anchor=tk.W, width=200, stretch=False)
tree.pack(expand=True, fill="both", padx=10, pady=20)
tree.bind('<ButtonRelease-1>', lambda e: entry_value.focus())

# Log view for displaying the 5 most recent changes
log_view_frame = tk.Frame(root, height=100)
log_view_frame.pack(side=tk.BOTTOM, fill='x', padx=10, pady=10)

log_view_columns = ("Timestamp", "Item", "Action")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", height=5)
for col in log_view_columns:
    log_view.heading(col, text=col, anchor=tk.W)
    log_view.column(col, anchor=tk.W, width=150, stretch=False)
log_view.pack(side='bottom', fill='x')

# Start the GUI event loop
root.after(100, update_treeview)
root.mainloop()
