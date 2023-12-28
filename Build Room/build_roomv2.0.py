import tkinter as tk
from tkinter import ttk
from pathlib import Path
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Define a class for the application
class InventoryApp(tk.Tk):
    def __init__(self, workbook_path):
        super().__init__()
        # ... other initialization code ...
        self.create_widgets()
        self.populate_treeview(self.treeview_items, '4.2 Items')  # This should be after create_widgets
        self.populate_treeview(self.treeview_timestamps, '4.2 Timestamps')


    def load_or_create_workbook(self):
        if Path(self.workbook_path).exists():
            self.workbook = load_workbook(self.workbook_path)
        else:
            self.workbook = Workbook()
            self.create_sheets_with_headers()

    def create_sheets_with_headers(self):
        sheets_headers = {
            '4.2 Items': ["Item", "LastCount", "NewCount"],
            '4.2 Timestamps': ["Timestamp", "Item", "Action", "SAN Number"],
            'BR Items': ["Item", "LastCount", "NewCount"],
            'BR Timestamps': ["Timestamp", "Item", "Action", "SAN Number"],
            'Project Designated Items': ["Item", "LastCount", "NewCount"],
            'Project Designated Timestamps': ["Timestamp", "Item", "Action", "SAN Number"],
            'All SANs': ["Item", "SAN Number", "Timestamp"]
        }

        self.workbook.active.title = '4.2 Items'
        for sheet_name, headers in sheets_headers.items():
            if sheet_name not in self.workbook.sheetnames:
                self.workbook.create_sheet(sheet_name)
            sheet = self.workbook[sheet_name]
            sheet.append(headers)
        
        self.workbook.save(self.workbook_path)

def create_widgets(self):
        # ... code to create widgets including treeviews ...
        self.treeview_items = self.create_treeview()
        self.treeview_timestamps = self.create_treeview()



        self.btn_basement = ttk.Button(self.buttons_frame, text="Basement 4.2",
                                       command=lambda: self.switch_sheet('4.2 Items', '4.2 Timestamps'))
        self.btn_basement.pack(side='left')

        self.btn_build_room = ttk.Button(self.buttons_frame, text="Build Room",
                                         command=lambda: self.switch_sheet('BR Items', 'BR Timestamps'))
        self.btn_build_room.pack(side='left')

        self.entry_field = tk.Entry(self.buttons_frame, validate='key',
                                    validatecommand=(self.register(self.validate_number_input), '%P'))
        self.entry_field.pack(side='left')

        self.btn_minus = ttk.Button(self.buttons_frame, text="-",
                                    command=lambda: self.update_count(-1))
        self.btn_minus.pack(side='left')

        self.btn_plus = ttk.Button(self.buttons_frame, text="+",
                                   command=lambda: self.update_count(1))
        self.btn_plus.pack(side='left')

        self.btn_export = ttk.Button(self.buttons_frame, text=".xlsx",
                                     command=self.export_to_excel)
        self.btn_export.pack(side='left')

        self.treeview_items = self.create_treeview()
        self.treeview_timestamps = self.create_treeview()

def validate_number_input(self, value_if_allowed):
    if value_if_allowed == '':
         return True
    try:
        int(value_if_allowed)
        return True
    except ValueError:
        return False

    def update_count(self, delta):
        selected_item = self.treeview_items.focus()
        if selected_item:
            current_count = self.treeview_items.item(selected_item, 'values')[2]
            new_count = int(current_count) + delta
            self.treeview_items.item(selected_item, values=(self.treeview_items.item(selected_item, 'values')[0],
                                                            current_count, new_count))

    def export_to_excel(self):
        # Placeholder for export functionality
        pass

    def create_treeview(self):
        treeview = ttk.Treeview(self)
        treeview.pack(fill='both', expand=True)
        return treeview

def populate_treeview(self, treeview, sheet_name):
        # Ensure this method definition is before it's called in __init__
        treeview.delete(*treeview.get_children())
        sheet = self.workbook[sheet_name]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        treeview["columns"] = headers
        treeview["show"] = "headings"  # Hide the first empty column

        for header in headers:
            treeview.heading(header, text=header)
            treeview.column(header, width=100, anchor='center')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            treeview.insert('', 'end', values=row)



def switch_sheet(self, items_sheet, timestamps_sheet):
    self.populate_treeview(self.treeview_items, items_sheet)
    self.populate_treeview(self.treeview_timestamps, timestamps_sheet)


if __name__ == "__main__":
    app = InventoryApp('EUC_Perth_Assets.xlsx')
    app.mainloop()
