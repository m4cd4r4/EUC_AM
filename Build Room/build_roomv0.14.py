import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl import Workbook
import tkinter.font as tkFont
from datetime import datetime
import os

# Initialize Tkinter root
root = tk.Tk()
root.title("Store-Room 4.2")
root.geometry("800x600")

# Define a large font for buttons
button_font = tkFont.Font(size=16)

# Load the workbook or create it if it doesn't exist
try:
    workbook = load_workbook('EUC_Build_Room.xlsx')
    sheet = workbook.active
    try:
        log_sheet = workbook['Sheet2']  # Attempt to access the second sheet
    except KeyError:
        log_sheet = workbook.create_sheet('Sheet2')  # Create the second sheet if it doesn't exist
        log_sheet.append(["Timestamp", "Item", "Action"])  # Create headers if not present
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Item", "LastCount", "NewCount"])  # Create headers if not present
    log_sheet = workbook.create_sheet('Sheet2')  # Create the second sheet
    log_sheet.append(["Timestamp", "Item", "Action"])  # Create headers if not present
    workbook.save('EUC_Build_Room.xlsx')

# Function to update the Treeview widget with the spreadsheet data
def update_treeview():
    for row in tree.get_children():
        tree.delete(row)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)

# Function to log the changes to the second sheet and update the log view
def log_change(item, action):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_sheet.append([timestamp, item, action])
    workbook.save('EUC_Build_Room.xlsx')
    update_log_view()

# Function to update the log view with the 5 most recent changes
def update_log_view():
    for row in log_view.get_children():
        log_view.delete(row)
    for row in log_sheet.iter_rows(min_row=max(1, log_sheet.max_row - 4), max_row=log_sheet.max_row, values_only=True):
        log_view.insert('', 'end', values=row)

# Function to update counts in the spreadsheet
def update_count(operation):
    try:
        selected_item = tree.item(tree.focus())['values'][0]  # Get the selected item from the treeview
        if not selected_item:
            print("No item selected.")
            return
        input_value = int(entry_value.get())
        item_found = False
        for row in sheet.iter_rows(min_row=2, values_only=False):
            if row[0].value == selected_item:
                item_found = True
                # Update LastCount with the value from NewCount
                row[1].value = row[2].value
                # Perform the operation
                new_value = (row[2].value or 0) + input_value if operation == 'add' else (row[2].value or 0) - input_value
                row[2].value = new_value
                # Save the workbook
                workbook.save('EUC_Build_Room.xlsx')
                print(f"Updated '{selected_item}' with new count {new_value}")
                update_treeview()  # Update the Treeview with new data
                # Log the change
                action = 'Add' if operation == 'add' else 'Subtract'
                log_change(selected_item, f"{action} {input_value}")
                break
        if not item_found:
            print(f"Item '{selected_item}' not found in the spreadsheet.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Create a frame to hold the widgets
frame = tk.Frame(root)
frame.pack(padx=10, pady=10, fill='both', expand=True)

# Entry field and buttons layout
entry_frame = tk.Frame(frame)
entry_frame.pack(pady=10)

# Subtract Button
button_subtract = tk.Button(entry_frame, text="-", command=lambda: update_count('subtract'), font=button_font)
button_subtract.pack(side=tk.LEFT, padx=5)

# Entry Value
entry_value = tk.Entry(entry_frame, width=10, font=button_font)
entry_value.pack(side=tk.LEFT)

# Add Button
button_add = tk.Button(entry_frame, text="+", command=lambda: update_count('add'), font=button_font)
button_add.pack(side=tk.LEFT, padx=5)

# Treeview for displaying spreadsheet data
columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', height=10)
tree.heading("Item", text="Item", anchor=tk.W)  # Align "Item" text to the left
tree.column("Item", anchor=tk.W, stretch=False)
tree.column("LastCount", anchor=tk.CENTER, stretch=False)
tree.column("NewCount", anchor=tk.CENTER, stretch=False)
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=200, minwidth=50, stretch=False)  # Set minimum width and disable stretching
tree.pack(expand=True, fill="both", padx=10, pady=10)
tree.bind('<ButtonRelease-1>', lambda e: entry_value.focus())  # Focus on entry field when an item is selected

# Log view for displaying the 5 most recent changes
log_view_frame = tk.Frame(root, height=100)
log_view_frame.pack(side=tk.BOTTOM, fill='x', padx=10, pady=10)

log_view_columns = ("Timestamp", "Item", "Action")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="tree", height=5)
for col in log_view_columns:
    log_view.column(col, width=150, anchor=tk.CENTER)
log_view.pack(side='bottom', fill='x')

# Load the logo image and place it at the top right corner
logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'santos.png')
logo_image = tk.PhotoImage(file=logo_path)
# Resize the logo by reducing it by 80%
logo_image = logo_image.subsample(5, 5)  # subsample by a factor of 5 to reduce the size by 80%
logo_label = tk.Label(root, image=logo_image)
logo_label.pack(side=tk.TOP, anchor='ne', padx=10, pady=10)

# Initially populate the Treeview and the Log View
update_treeview()
update_log_view()

# Start the GUI event loop
root.mainloop()
