import customtkinter as ctk
from openpyxl import load_workbook, Workbook
from datetime import datetime
import os
import tkinter as tk
from tkinter import ttk, simpledialog

# Initialize Tkinter root with CustomTkinter
root = ctk.CTk()
root.title("Store-Room 4.2")
root.geometry("800x600")

# Load the workbook or create it if it doesn't exist
workbook_path = 'EUC_Perth_Assets.xlsx'
workbook = load_workbook(workbook_path) if os.path.exists(workbook_path) else Workbook()

# Check for existing sheets or create them
sheet_names = workbook.sheetnames
sheet = workbook['4.2_Items'] if '4.2_Items' in sheet_names else workbook.create_sheet('4.2_Items')
log_sheet = workbook['4.2_Timestamps'] if '4.2_Timestamps' in sheet_names else workbook.create_sheet('4.2_Timestamps')
br_sheet = workbook['BR_Items'] if 'BR_Items' in sheet_names else workbook.create_sheet('BR_Items')
br_log_sheet = workbook['BR_Timestamps'] if 'BR_Timestamps' in sheet_names else workbook.create_sheet('BR_Timestamps')

# Initialize headers if the workbook was just created
if not os.path.exists(workbook_path):
    sheet.append(["Item", "LastCount", "NewCount"])
    log_sheet.append(["Timestamp", "Item", "Action", "SAN Number"])
    br_sheet.append(["Item", "LastCount", "NewCount"])
    br_log_sheet.append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook.save(workbook_path)

# Define a larger font for Treeview
style = ttk.Style()
larger_font = ('Helvetica', 12)  # Increase font size by 50%
style.configure("Treeview", font=larger_font)

# Custom dialog class for SAN input
class SANInputDialog(simpledialog.Dialog):
    def body(self, master):
        self.entry = ctk.CTkEntry(master, width=200, justify='center', font=("Helvetica", 16))
        self.entry.pack(pady=10)
        self.create_numpad(master)
        return self.entry

    def create_numpad(self, master):
        buttons = [
            ('1', 1), ('2', 2), ('3', 3),
            ('4', 4), ('5', 5), ('6', 6),
            ('7', 7), ('8', 8), ('9', 9),
            ('Del', 10), ('0', 11), ('Enter', 12)
        ]
        for btn_text, index in buttons:
            action = lambda val=btn_text: self.on_numpad_click(val)
            btn = ctk.CTkButton(master, text=btn_text, command=action, width=10, height=2)
            btn.grid(row=index//3, column=index%3)

    def on_numpad_click(self, value):
        if value == 'Del':
            current = self.entry.get()
            self.entry.delete(0, tk.END)
            self.entry.insert(0, current[:-1])
        elif value == 'Enter':
            self.result = self.entry.get()
            self.destroy()
        else:
            self.entry.insert(tk.END, value)

# Function to show SAN input dialog
def show_san_input():
    dialog = SANInputDialog(root, "Enter SAN Number")
    return dialog.result

# Function to update the Treeview widget with the spreadsheet data
def update_treeview():
    tree.delete(*tree.get_children())  # Clear the existing treeview entries
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', 'end', values=row)

# Function to log the changes to the log sheet and update the log view
def log_change(item, action, target_sheet, san_number=""):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    target_sheet.append([timestamp, item, action, san_number])
    workbook.save(workbook_path)
    update_log_view()

# Function to update the log view with the 5 most recent changes in ascending order
def update_log_view():
    log_view.delete(*log_view.get_children())  # Clear existing entries
    all_rows = list(log_sheet.iter_rows(min_row=2, max_row=log_sheet.max_row, values_only=True))
    sorted_rows = sorted(all_rows, key=lambda r: r[0] if r[0] else datetime.min)
    for row in sorted_rows[-5:]:
        log_view.insert('', 0, values=row)  # Inserting at index 0 to put the most recent at the top

# Function to update counts in the spreadsheet
def update_count(operation):
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if selected_item:
        try:
            input_value = int(entry_value.get())
            if operation == 'subtract':
                san_number = show_san_input()
                target_sheet = log_sheet if selected_item in sheet.iter_rows(min_row=2, values_only=True) else br_log_sheet
                log_change(selected_item, f"{operation.capitalize()} {input_value}", target_sheet, san_number)
            else:
                log_change(selected_item, f"{operation.capitalize()} {input_value}", log_sheet)
            workbook.save(workbook_path)
            update_treeview()
        except ValueError as e:
            print(f"Invalid input for count update: {e}")

# Create a frame to hold the widgets using CustomTkinter
frame = ctk.CTkFrame(root)
frame.pack(padx=10, pady=10, fill='both', expand=True)

# Entry field and buttons layout using CustomTkinter
entry_frame = ctk.CTkFrame(frame)
entry_frame.pack(pady=10)

button_subtract = ctk.CTkButton(entry_frame, text="-", command=lambda: update_count('subtract'), font=("Helvetica", 16))
button_subtract.pack(side='left', padx=5)

entry_value = tk.Entry(entry_frame, font=("Helvetica", 16), justify='center', width=10)
entry_value.pack(side='left')

button_add = ctk.CTkButton(entry_frame, text="+", command=lambda: update_count('add'), font=("Helvetica", 16))
button_add.pack(side='left', padx=5)

columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', height=8, style="Treeview")
for col in columns:
    tree.heading(col, text=col, anchor=tk.W)
    tree.column(col, anchor=tk.W, width=200, stretch=False)
tree.pack(expand=True, fill="both", padx=10, pady=20)
tree.bind('<ButtonRelease-1>', lambda e: entry_value.focus())

# Log view for displaying the 5 most recent changes
log_view_frame = ctk.CTkFrame(root)
log_view_frame.pack(side=tk.BOTTOM, fill='x', padx=10, pady=10)

log_view_columns = ("Timestamp", "Item", "Action", "SAN Number")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", height=5, style="Treeview")
for col in log_view_columns:
    log_view.heading(col, text=col, anchor=tk.W)
    log_view.column(col, anchor=tk.W, width=150, stretch=False)
log_view.pack(side='bottom', fill='x')

# Start the GUI event loop
root.after(100, update_treeview)
root.mainloop()
