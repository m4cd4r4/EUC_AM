# Add dark-mode
# Build Room\build_roomv2.py
# Author: Macdara o Murchu
# 31.12.23

import logging.config
from pathlib import Path
from tkinter import Menu
import customtkinter as ctk
import os
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from datetime import datetime
import subprocess

logging_conf_path = Path('logging.conf')
if logging_conf_path.exists() and logging_conf_path.stat().st_size > 0:
    try:
        logging.config.fileConfig(logging_conf_path)
    except Exception as e:
        logging.error(f"Error configuring logging: {e}", exc_info=True)
else:
    logging.basicConfig(level=logging.DEBUG)

root = ctk.CTk()
root.title("Perth EUC Assets")
root.geometry("550x600")

script_directory = Path(__file__).parent
workbook_path = script_directory / 'EUC_Perth_Assets.xlsx'
if Path(workbook_path).exists():
    workbook = load_workbook(workbook_path)
else:
    workbook = Workbook()
    workbook.active.title = '4.2 Items'
    workbook.create_sheet('4.2 Timestamps')
    workbook.create_sheet('BR Items')
    workbook.create_sheet('BR Timestamps')
    workbook.create_sheet('Project Designated Items')
    workbook.create_sheet('Project Designated Timestamps')
    workbook.create_sheet('All SANs')
    workbook['4.2 Items'].append(["Item", "LastCount", "NewCount"])
    workbook['4.2 Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['BR Items'].append(["Item", "LastCount", "NewCount"])
    workbook['BR Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['Project Designated Items'].append(["Item", "LastCount", "NewCount"])
    workbook['Project Designated Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['All SANs'].append(["SAN Number", "Item", "Timestamp"])
    workbook.save(workbook_path)

all_sans_sheet = workbook['All SANs']
sheets = {'original': ('4.2 Items', '4.2 Timestamps'), 'backup': ('BR Items', 'BR Timestamps')}
current_sheets = sheets['original']

style = ttk.Style()
style.configure("Treeview", font=('Helvetica', 12))

vcmd = (root.register(lambda P: P.isdigit() or P == ""), '%P')

class SANInputDialog(tk.Toplevel):
    def __init__(self, parent, title=None):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.parent = parent
        self.result = None
        self.create_widgets()
        self.grab_set()
        self.geometry(f"+{parent.winfo_rootx() + parent.winfo_width() // 2 - 100}+{parent.winfo_rooty() + parent.winfo_height() // 2 - 50}")
        self.wait_window(self)

    def create_widgets(self):
        self.entry = ttk.Entry(self, validate="key", validatecommand=vcmd)
        self.entry.pack(padx=5, pady=5)
        button_frame = tk.Frame(self)
        button_frame.pack(pady=5)
        submit_button = ttk.Button(button_frame, text="Submit", command=self.on_submit)
        submit_button.pack(side='left', padx=5)
        cancel_button = ttk.Button(button_frame, text="Cancel", command=self.on_cancel)
        cancel_button.pack(side='left', padx=5)

    def on_submit(self):
        san_input = self.entry.get()
        if san_input and len(san_input) >= 5 and len(san_input) <= 6:
            self.result = san_input
            self.destroy()
        else:
            tk.messagebox.showerror("Error", "Please enter a valid SAN number.", parent=self)
            self.entry.focus_set()

    def on_cancel(self):
        self.result = None
        self.destroy()

def is_san_unique(san_number):
    return all(san_number != row[1] for row in all_sans_sheet.iter_rows(min_row=2, values_only=True))

def show_san_input():
    dialog = SANInputDialog(root, "Enter SAN Number")
    return dialog.result

def open_spreadsheet():
    try:
        if os.name == 'nt':
            os.startfile(workbook_path)
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.run([opener, workbook_path])
    except Exception as e:
        tk.messagebox.showerror("Error", f"Failed to open the spreadsheet: {e}")

frame = ctk.CTkFrame(root)
frame.pack(padx=10, pady=10, fill='both', expand=True)
entry_frame = ctk.CTkFrame(frame)
entry_frame.pack(pady=10)

button_width = 50
button_1 = ctk.CTkButton(entry_frame, text="Basement 4.2", command=lambda: switch_sheets('original'), width=button_width, font=("Helvetica", 16))
button_1.pack(side='left', padx=5)
button_2 = ctk.CTkButton(entry_frame, text="Build Room", command=lambda: switch_sheets('backup'), width=button_width, font=("Helvetica", 16))
button_2.pack(side='left', padx=5)
button_subtract = ctk.CTkButton(entry_frame, text="-", command=lambda: update_count('subtract'), width=button_width, font=("Helvetica", 16))
button_subtract.pack(side='left', padx=5)
entry_value = tk.Entry(entry_frame, font=("Helvetica", 16), justify='center', width=10, validate="key", validatecommand=vcmd)
entry_value.pack(side='left')
button_add = ctk.CTkButton(entry_frame, text="+", command=lambda: update_count('add'), width=button_width, font=("Helvetica", 16))
button_add.pack(side='left', padx=5)
xlsx_button = ctk.CTkButton(entry_frame, text=".xlsx", command=open_spreadsheet, width=button_width, font=("Helvetica", 16))
xlsx_button.pack(side='left', padx=5)

# current_mode = "light"  # Global variable to track the current mode
def apply_treeview_style(style_name):
    style = ttk.Style()
    style.configure(style_name + ".Treeview", font=('Helvetica', 12, 'bold'), foreground="white")
    style.configure(style_name + ".Treeview.Heading", font=('Helvetica', 13, 'bold'))

def update_treeview_colors():
    style = ttk.Style()
    if ctk.get_appearance_mode() == "dark":
        style.configure("Treeview", background="darkgrey", fieldbackground="darkgrey", foreground="white")
        style.map("Treeview", background=[('selected', 'darkblue')], foreground=[('selected', 'white')])
        style.configure("Treeview.Heading", background="grey", foreground="white")
        tree.tag_configure('oddrow', background='#505050')  # Dark grey
        tree.tag_configure('evenrow', background='#404040')  # Lighter grey
        log_view.tag_configure('oddrow', background='#505050')
        log_view.tag_configure('evenrow', background='#404040')
    else:
        style.configure("Treeview", background="white", fieldbackground="white", foreground="black")
        style.map("Treeview", background=[('selected', 'blue')], foreground=[('selected', 'white')])
        style.configure("Treeview.Heading", background="lightgrey", foreground="black")
        tree.tag_configure('oddrow', background='#f0f0f0')  # Light grey
        tree.tag_configure('evenrow', background='white')
        log_view.tag_configure('oddrow', background='#f0f0f0')
        log_view.tag_configure('evenrow', background='white')

def toggle_theme():
    global tree, log_view
    current_mode = ctk.get_appearance_mode()
    new_mode = "dark" if current_mode == "light" else "light"
    ctk.set_appearance_mode(new_mode)
    update_treeview_colors()

# Initialize the treeview with default light mode colors
update_treeview_colors()

# Create a menu bar
menu_bar = tk.Menu(root)
root.config(menu=menu_bar)

# Create a "File" menu
file_menu = tk.Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="File", menu=file_menu)

# Add "Toggle Dark Mode" to the "File" menu
file_menu.add_command(label="Toggle Dark Mode", command=toggle_theme)

def update_treeview():
    tree.delete(*tree.get_children())
    workbook = load_workbook(workbook_path)
    item_sheet = workbook[current_sheets[0]]
    row_count = 0
    for row in item_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            if row_count % 2 == 0:
                bg_color = 'white'
            else:
                bg_color = '#f0f0f0'
            tree.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
            tree.tag_configure('oddrow', background='#f0f0f0')
            tree.tag_configure('evenrow', background='white')
            row_count += 1

def log_change(item, action, san_number="", timestamp_sheet=None):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        if timestamp_sheet is not None:
            san_number = f"SAN{san_number}" if san_number and not san_number.startswith('SAN') else san_number
            timestamp_sheet.append([timestamp, item, action, san_number])
            workbook.save(workbook_path)
            update_log_view()
            logging.info(f"Logged change: Time: {timestamp}, Item: {item}, Action: {action}, SAN: {san_number}")
        else:
            logging.error("No timestamp sheet provided for logging.")
    except Exception as e:
        logging.error(f"Failed to log change: {e}")
        tk.messagebox.showerror("Error", f"Failed to log change: {e}")

def switch_sheets(sheet_type):
    global current_sheets
    current_sheets = sheets[sheet_type]
    update_treeview()
    update_log_view()

def update_log_view():
    if 'log_view' in globals():
        log_view.delete(*log_view.get_children())
        log_sheet = workbook[current_sheets[1]]
        all_rows = list(log_sheet.iter_rows(min_row=2, values_only=True))
        # Adjust the sorting to use the first column (timestamp)
        sorted_rows = sorted(all_rows, key=lambda r: datetime.strptime(r[0], "%Y-%m-%d %H:%M:%S") if r[0] else datetime.min, reverse=True)
        row_count = 0
        for row in sorted_rows:
            if row[0] is not None:
                log_view.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
                log_view.tag_configure('oddrow', background='#f0f0f0')
                log_view.tag_configure('evenrow', background='white')
                row_count += 1

def update_count(operation):
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if selected_item:
        input_value = entry_value.get()
        if input_value.isdigit():
            input_value = int(input_value)
            item_sheet = workbook[current_sheets[0]]
            timestamp_sheet = workbook[current_sheets[1]]
            san_required = any(g in selected_item for g in ["G8", "G9", "G10"])

            if san_required:
                san_count = 0
                for _ in range(abs(input_value)):
                    while True:
                        san_number = show_san_input()
                        if san_number is None:  # User cancelled the input
                            return
                        san_number = "SAN" + san_number

                        if operation == 'add':
                            if is_san_unique(san_number):
                                all_sans_sheet.append([selected_item, san_number, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                                log_change(selected_item, operation, san_number, timestamp_sheet)
                                san_count += 1
                                break
                            else:
                                tk.messagebox.showerror("Error", f"Duplicate or already used SAN number: {san_number}", parent=root)
                        elif operation == 'subtract':
                            if any(san_number == row[1] for row in all_sans_sheet.iter_rows(min_row=2, values_only=True)):
                                for row in all_sans_sheet.iter_rows(min_row=2):
                                    if row[1].value == san_number:
                                        all_sans_sheet.delete_rows(row[0].row)
                                        log_change(selected_item, operation, san_number, timestamp_sheet)
                                        san_count += 1
                                        break
                                break
                            else:
                                tk.messagebox.showerror("Error", f"That SAN number does not exist in the All SANs sheet: {san_number}", parent=root)
                                continue
                input_value = san_count  # Adjust input_value to the number of valid SANs processed

            # Adjust item counts
            for row in item_sheet.iter_rows(min_row=2):
                if row[0].value == selected_item:
                    row[1].value = row[2].value or 0
                    if operation == 'add':
                        row[2].value = (row[2].value or 0) + input_value
                    elif operation == 'subtract':
                        row[2].value = max((row[2].value or 0) - input_value, 0)

            # Log the change for items not requiring SAN
            if not san_required:
                log_change(selected_item, operation, "", timestamp_sheet)

            workbook.save(workbook_path)
            update_treeview()
            update_log_view()

columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', style="Treeview")
for col in columns:
    tree.heading(col, text=col, anchor='w')
    tree.column(col, anchor='w', width=200, stretch=False)
tree.pack(expand=True, fill="both", padx=10, pady=20)

# Define and set up the 'log_view' widget
log_view_frame = ctk.CTkFrame(root)
log_view_frame.pack(side=tk.BOTTOM, fill='both', expand=True, padx=10, pady=10)
log_view_columns = ("Timestamp", "Item", "Action", "SAN Number")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", style="Treeview", height=8)
for col in log_view_columns:
    log_view.heading(col, text=col, anchor='w')
    log_view.column(col, anchor='w', width=170, stretch=False)
scrollbar_log = ttk.Scrollbar(log_view_frame, orient="vertical", command=log_view.yview)
scrollbar_log.pack(side='right', fill='y')
log_view.configure(yscrollcommand=scrollbar_log.set)
log_view.pack(expand=True, fill='both')

# Call the function to update the Treeview colors
update_treeview_colors()

root.mainloop()
