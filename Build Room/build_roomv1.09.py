# Add shading to data in logview & treeview, for improved readability
# Build Room\build_roomv1.07.py
# Author: Macdara o Murchu

import logging.config
from pathlib import Path
import customtkinter as ctk
import os
import tkinter as tk
from tkinter import simpledialog, ttk
from openpyxl import load_workbook, Workbook
from datetime import datetime
import subprocess  # Import subprocess for non-Windows systems
import logging

# Path to the logging configuration file
logging_conf_path = Path('logging.conf')

# Check if logging configuration file exists and is not empty
if logging_conf_path.exists() and logging_conf_path.stat().st_size > 0:
    try:
        logging.config.fileConfig(logging_conf_path)
    except Exception as e:
        print(f"Error configuring logging: {e}")
else:
    print("Logging configuration file is missing or empty. Using default logging settings.")
    logging.basicConfig(level=logging.DEBUG)

# Configure logging - Disable this for debugging of empty logging.conf file
logging.config.fileConfig('logging.conf')

root = ctk.CTk()
root.title("Perth EUC Assets")
root.geometry("550x600")

# Load the workbook or create it if it doesn't exist
script_directory = Path(__file__).parent  # Get the parent directory of the script
workbook_path = script_directory / 'EUC_Perth_Assets.xlsx'  # Construct the path to the workbook in the same directory as the script
if Path(workbook_path).exists():
    workbook = load_workbook(workbook_path)
else:
    workbook = Workbook()
    workbook.active.title = '4.2 Items'
    workbook.create_sheet('4.2 Timestamps')
    workbook.create_sheet('BR Items')
    workbook.create_sheet('BR Timestamps')
    workbook.create_sheet('Project Designated Items')
    workbook.create_sheet('Project Designated Timestamps')
    workbook.create_sheet('All SANs')
    workbook['4.2 Items'].append(["Item", "LastCount", "NewCount"])
    workbook['4.2 Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['BR Items'].append(["Item", "LastCount", "NewCount"])
    workbook['BR Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['Project Designated Items'].append(["Item", "LastCount", "NewCount"])
    workbook['Project Designated Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['All SANs'].append(["Item", "SAN Number", "Timestamp"])
    workbook.save(workbook_path)

# Sheets reference
sheets = {
    'original': ('4.2 Items', '4.2 Timestamps'),
    'backup': ('BR Items', 'BR Timestamps')
}

# Currently active sheets
current_sheets = sheets['original']

# Define a larger font for Treeview
style = ttk.Style()
larger_font = ('Helvetica', 12)  # Increase font size by 50%
style.configure("Treeview", font=larger_font)

# Custom validation function to allow only numeric input
def validate_input(P):
    # Check if the input is either empty (to allow clearing the field) or a number
    return P.isdigit() or P == ""

# Configure the Tkinter validation command for numbers only entry
vcmd = (root.register(validate_input), '%P')


# Custom dialog class for SAN input
class SANInputDialog(tk.Toplevel):
    def __init__(self, parent, title=None):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.parent = parent
        self.result = None
        self.create_widgets()
        self.grab_set()

        # Center the dialog on the parent window
        self.geometry(f"+{parent.winfo_rootx() + parent.winfo_width() // 2 - 100}+{parent.winfo_rooty() + parent.winfo_height() // 2 - 50}")

        self.wait_window(self)

    def create_widgets(self):
        self.entry = ttk.Entry(self, validate="key", validatecommand=vcmd)
        self.entry.pack(padx=5, pady=5)
        button_frame = tk.Frame(self)
        button_frame.pack(pady=5)

        submit_button = ttk.Button(button_frame, text="Submit", command=self.on_submit)
        submit_button.pack(side='left', padx=5)

        cancel_button = ttk.Button(button_frame, text="Cancel", command=self.on_cancel)
        cancel_button.pack(side='left', padx=5)

    def on_submit(self):
        san_input = self.entry.get()
        if san_input and len(san_input) >= 5 and len(san_input) <= 6:
            self.result = san_input
            self.destroy()
        else:
            tk.messagebox.showerror("Error", "Please enter a valid SAN with at least 5 characters.", parent=self)
            self.entry.focus_set()

    def on_cancel(self):
        self.result = None
        self.destroy()


# Define the function just after the SANInputDialog class
def is_san_unique(san_number):
    all_sans_sheet = workbook['All SANs']
    for row in all_sans_sheet.iter_rows(min_row=2, values_only=True):
        if san_number == row[1]:  # Assuming the SAN Number is in the second column
            return False
    return True

def show_san_input():
    dialog = SANInputDialog(root, "Enter SAN Number")
    return dialog.result


# Custom validation function to allow only numeric input
def validate_input(P):
    # Check if the input is either empty (to allow clearing the field) or a number
    return P.isdigit() or P == ""

# Function to open the spreadsheet
def open_spreadsheet():
    try:
        if os.name == 'nt':  # Windows
            os.startfile(workbook_path)
        else:  # macOS, Linux, etc.
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.run([opener, workbook_path])
    except Exception as e:
        tk.messagebox.showerror("Error", f"Failed to open the spreadsheet: {e}")

# Configure the Tkinter validation command
vcmd = (root.register(validate_input), '%P')

# Create a frame to hold the widgets using CustomTkinter
frame = ctk.CTkFrame(root)
frame.pack(padx=10, pady=10, fill='both', expand=True)

# Entry field and buttons layout using CustomTkinter
entry_frame = ctk.CTkFrame(frame)
entry_frame.pack(pady=10)

# Button width calculation
button_width = 50  # Reduced width by 66%

# Sheet switch buttons
button_1 = ctk.CTkButton(entry_frame, text="Basement 4.2", command=lambda: switch_sheets('original'), width=button_width, font=("Helvetica", 16))
button_1.pack(side='left', padx=5)

button_2 = ctk.CTkButton(entry_frame, text="Build Room", command=lambda: switch_sheets('backup'), width=button_width, font=("Helvetica", 16))
button_2.pack(side='left', padx=5)

# Subtract button
button_subtract = ctk.CTkButton(entry_frame, text="-", command=lambda: update_count('subtract'), width=button_width, font=("Helvetica", 16))
button_subtract.pack(side='left', padx=5)

# Entry for count update with validation applied
entry_value = tk.Entry(entry_frame, font=("Helvetica", 16), justify='center', width=10, validate="key", validatecommand=vcmd)
entry_value.pack(side='left')

# Add button
button_add = ctk.CTkButton(entry_frame, text="+", command=lambda: update_count('add'), width=button_width, font=("Helvetica", 16))
button_add.pack(side='left', padx=5)

# .xlsx button
xlsx_button = ctk.CTkButton(entry_frame, text=".xlsx", command=open_spreadsheet, width=button_width, font=("Helvetica", 16))
xlsx_button.pack(side='left', padx=5)


# Function to update the Treeview widget with the spreadsheet data
def update_treeview():
    tree.delete(*tree.get_children())  # Clear the existing treeview entries
    # Open the workbook again to read the latest data
    workbook = load_workbook(workbook_path)
    item_sheet = workbook[current_sheets[0]]
    
    # Track the row count to alternate row colors
    row_count = 0
    
    for row in item_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:  # Ensure that the first cell in the row is not empty
            # Determine the background color based on the row count
            if row_count % 2 == 0:
                bg_color = 'white'  # Use white for even rows
            else:
                bg_color = '#f0f0f0'  # Use a light gray color for odd rows
            
            # Insert the row with the specified background color
            tree.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
            
            # Apply the background color to the inserted row
            tree.tag_configure('oddrow', background='#f0f0f0')
            tree.tag_configure('evenrow', background='white')
            
            row_count += 1

# Function to log the changes to the log sheet and update the log view
def log_change(item, action, san_number="", timestamp_sheet=None):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        if timestamp_sheet is not None:
            # Prepend 'SAN' to the san_number if it's not empty and doesn't already start with 'SAN'
            san_number = f"SAN{san_number}" if san_number and not san_number.startswith('SAN') else san_number

            # Assuming the order of columns in the timestamp sheet is: Item, Action, SAN Number, Time
            timestamp_sheet.append([item, action, san_number, timestamp])
            workbook.save(workbook_path)
            logging.info(f"Logged change: Item: {item}, Action: {action}, SAN: {san_number}, Time: {timestamp}")
        else:
            logging.error("No timestamp sheet provided for logging.")
    except Exception as e:
        logging.error(f"Failed to log change: {e}")
        tk.messagebox.showerror("Error", f"Failed to log change: {e}")

# Function to update the Log View with sorted and reordered data
def update_log_view():
    if 'log_view' in globals():
        log_view.delete(*log_view.get_children())  # Clear existing entries
        log_sheet = workbook[current_sheets[1]]
        
        # Retrieve all rows from the log sheet, skipping the header
        all_rows = list(log_sheet.iter_rows(min_row=2, values_only=True))
        
        # Sort the rows by the timestamp column in descending order to display newest-first
        sorted_rows = sorted(all_rows, key=lambda r: r[0] or '', reverse=True)
        
        # Track the row count to alternate row colors
        row_count = 0
        
        # Insert the sorted rows into the Treeview with alternating row colors
        for row in sorted_rows:
            if row[0] is not None:  # Only insert rows with a timestamp
                # Determine the background color based on the row count
                if row_count % 2 == 0:
                    bg_color = 'white'  # Use white for even rows
                else:
                    bg_color = '#f0f0f0'  # Use a light gray color for odd rows
                
                # Insert the row with the specified background color
                log_view.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
                
                # Apply the background color to the inserted row
                log_view.tag_configure('oddrow', background='#f0f0f0')
                log_view.tag_configure('evenrow', background='white')
                
                row_count += 1
# Function to switch between original and backup sheets
def switch_sheets(sheet_type):
    global current_sheets
    current_sheets = sheets[sheet_type]
    update_treeview()
    update_log_view()

# Function to check for SAN number uniqueness in 'All SANs' sheet
def is_san_unique(san_number):
    all_sans_sheet = workbook['All SANs']
    for row in all_sans_sheet.iter_rows(min_row=2, values_only=True):
        if san_number == row[1]:  # Assuming the SAN Number is in the second column
            return False
    return True


def update_count(operation):
    try:
        selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
        if selected_item:
            input_value = int(entry_value.get())
            item_sheet = workbook[current_sheets[0]]
            timestamp_sheet = workbook[current_sheets[1]]

            # Update the item sheet only after successful SAN input
            updated_rows = []  # Keep track of the rows to be updated

            # Loop for getting SAN input and checking for uniqueness
            for _ in range(abs(input_value)):
                while True:  # Loop until a unique SAN is entered or the user cancels
                    san_number = show_san_input()
                    if san_number is None:  # If the input was cancelled, exit the update function immediately
                        logging.info(f"SAN input was canceled for item {selected_item}.")
                        return
                    san_number = "SAN" + san_number  # Prepend "SAN"
                    if is_san_unique(san_number):  # Check for uniqueness
                        all_sans_sheet = workbook['All SANs']
                        all_sans_sheet.append([selected_item, san_number, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                        # Log the change and break out of the while loop to process the next input
                        log_change(selected_item, operation, san_number, timestamp_sheet)
                        updated_rows.append(selected_item)  # Add the item to the list of items to update
                        break
                    else:
                        tk.messagebox.showerror("Duplicate SAN", "This SAN number is already recorded in the system. Please reenter a unique SAN.")
                        # The loop will continue, prompting the user again for a unique SAN

            # Now, update the counts for the rows that had unique SANs entered
            for row in item_sheet.iter_rows(min_row=2):
                if row[0].value in updated_rows:
                    row[1].value = row[2].value or 0
                    if operation == 'add':
                        row[2].value = (row[2].value or 0) + input_value
                    elif operation == 'subtract':
                        row[2].value = max((row[2].value or 0) - input_value, 0)

    except ValueError as e:
        logging.error(f"Invalid input for count update: {e}")
        tk.messagebox.showerror("Error", f"Invalid input for count update: {e}")

    finally:
        # Save the workbook and refresh the Treeview only if the update was successful
        workbook.save(workbook_path)
        update_treeview()


# Treeview for item display
columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', style="Treeview")
for col in columns:
    tree.heading(col, text=col, anchor='w')
    tree.column(col, anchor='w', width=200, stretch=False)
tree.pack(expand=True, fill="both", padx=10, pady=20)

log_view_frame = ctk.CTkFrame(root)
log_view_frame.pack(side=tk.BOTTOM, fill='both', expand=True, padx=10, pady=10)

log_view_columns = ("Timestamp", "Item", "Action", "SAN Number")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", style="Treeview", height=8)

for col in log_view_columns:
    log_view.heading(col, text=col, anchor='w')
    log_view.column(col, anchor='w', width=150, stretch=False)

# Scrollbar for the Log View
scrollbar_log = ttk.Scrollbar(log_view_frame, orient="vertical", command=log_view.yview)
scrollbar_log.pack(side='right', fill='y')
log_view.configure(yscrollcommand=scrollbar_log.set)

log_view.pack(expand=True, fill='both')


# Function to update the Log View with sorted and reordered data
def update_log_view():
    if 'log_view' in globals():
        log_view.delete(*log_view.get_children())  # Clear existing entries
        log_sheet = workbook[current_sheets[1]]
        
        # Retrieve all rows from the log sheet, skipping the header
        all_rows = list(log_sheet.iter_rows(min_row=2, values_only=True))
        
        # Sort the rows by the timestamp column in descending order to display newest-first
        sorted_rows = sorted(all_rows, key=lambda r: r[0] or '', reverse=True)
        
        # Insert the sorted rows into the Treeview
        for row in sorted_rows:
            if row[0] is not None:  # Only insert rows with a timestamp
                log_view.insert('', 'end', values=row)

# Initialize the Treeview and Log View with data
root.after(100, update_treeview)
update_log_view()  # Populate the log view when the application starts

# Start the GUI event loop
root.after(100, update_treeview)
root.mainloop()
