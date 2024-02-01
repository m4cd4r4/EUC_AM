# Add compulsory S/N entry for subtraction of wired & wireless headsets
# Add optional servicenow ticket number entry for wired & wireless headsets
# * If servicenow ticket number input box cancelled, a "Notes" form appears, logging info on the removed headset
#
# Build Room\build_roomv2.10.py
# Author: Macdara O Murchu
# 01.02.24

import logging.config
from pathlib import Path
from tkinter import Menu
import customtkinter as ctk
import os
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from datetime import datetime
import subprocess
import re

logging_conf_path = Path('logging.conf')
if logging_conf_path.exists() and logging_conf_path.stat().st_size > 0:
    try:
        logging.config.fileConfig(logging_conf_path)
    except Exception as e:
        logging.error(f"Error configuring logging: {e}", exc_info=True)
else:
    logging.basicConfig(level=logging.DEBUG)

def run_inventory_script():
    script_path = script_directory / "inventory-levels_4.2v1.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_4.2v1.py' does not exist in the directory.")

def run_build_room_inventory_script():
    script_path = script_directory / "inventory-levels_BRv1.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_BRv1.py' does not exist in the directory.")

def run_combined_rooms_inventory_script():
    script_path = script_directory / "inventory-levels_combinedv1.1.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_combinedv1.py' does not exist in the directory.")

root = ctk.CTk()
root.title("Perth EUC Assets")
root.geometry("500x650")

# Function to create and update the Headsets log view
def show_headsets_log():
    headsets_window = tk.Toplevel(root)
    headsets_window.title("Headsets Log")
    headsets_window.geometry("600x400")

    headset_columns = ("Serial Number", "ServiceNow #", "Notes")
    headsets_tree = ttk.Treeview(headsets_window, columns=headset_columns, show="headings")
    for col in headset_columns:
        headsets_tree.heading(col, text=col)
        headsets_tree.column(col, anchor='w')

    # Populate the Treeview
    if 'Headsets' in workbook.sheetnames:
        headset_sheet = workbook['Headsets']
        for row in headset_sheet.iter_rows(min_row=2, values_only=True):
            headsets_tree.insert('', 'end', values=row)

    headsets_tree.pack(expand=True, fill='both', padx=10, pady=10)

menu_bar = tk.Menu(root)
plots_menu = tk.Menu(menu_bar, tearoff=0)
plots_menu.add_command(label="Basement 4.2 Inventory", command=run_inventory_script)
plots_menu.add_command(label="Build Room Inventory", command=run_build_room_inventory_script)
plots_menu.add_command(label="Combined Inventory", command=run_combined_rooms_inventory_script)
plots_menu.add_command(label="Headsets", command=show_headsets_log)
menu_bar.add_cascade(label="Data", menu=plots_menu)
root.config(menu=menu_bar)

script_directory = Path(__file__).parent
workbook_path = script_directory / 'EUC_Perth_Assets.xlsx'
if Path(workbook_path).exists():
    workbook = load_workbook(workbook_path)
else:
    workbook = Workbook()
    workbook.active.title = '4.2 Items'
    workbook.create_sheet('4.2 Timestamps')
    workbook.create_sheet('BR Items')
    workbook.create_sheet('BR Timestamps')
    workbook.create_sheet('Project Designated Items')
    workbook.create_sheet('Project Designated Timestamps')
    workbook.create_sheet('All SANs')
    workbook['4.2 Items'].append(["Item", "LastCount", "NewCount"])
    workbook['4.2 Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['BR Items'].append(["Item", "LastCount", "NewCount"])
    workbook['BR Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['Project Designated Items'].append(["Item", "LastCount", "NewCount"])
    workbook['Project Designated Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['All SANs'].append(["SAN Number", "Item", "Timestamp"])
    workbook.save(workbook_path)

all_sans_sheet = workbook['All SANs']
sheets = {'original': ('4.2 Items', '4.2 Timestamps'), 'backup': ('BR Items', 'BR Timestamps')}
current_sheets = sheets['original']

style = ttk.Style()
style.configure("Treeview", font=('Helvetica', 12))

vcmd = (root.register(lambda P: P.isdigit() or P == ""), '%P')


class SANInputDialog(tk.Toplevel):
    def __init__(self, parent, title=None):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.parent = parent
        self.result = None
        self.create_widgets()
        self.grab_set()
        self.geometry(f"+{parent.winfo_rootx() + parent.winfo_width() // 2 - 100}+{parent.winfo_rooty() + parent.winfo_height() // 2 - 50}")
        self.wait_window(self)

    def create_widgets(self):
        self.entry = ttk.Entry(self, validate="key", validatecommand=vcmd)
        self.entry.pack(padx=5, pady=5)
        button_frame = tk.Frame(self)
        button_frame.pack(pady=5)
        submit_button = ttk.Button(button_frame, text="Submit", command=self.on_submit)
        submit_button.pack(side='left', padx=5)
        cancel_button = ttk.Button(button_frame, text="Cancel", command=self.on_cancel)
        cancel_button.pack(side='left', padx=5)

    def on_submit(self):
        ticket_number = self.entry.get()
        prefix = self.prefix_var.get()
        if ticket_number and prefix:
            self.result = f"{prefix}{ticket_number}"  # Dash removed between prefix and ticket number
            self.destroy()
        else:
            tk.messagebox.showerror("Error", "Please enter a valid Ticket Number.", parent=self)


    def on_cancel(self):
        self.result = None
        self.destroy()


def is_san_unique(san_number):
    # Adjust the search to account for the 'SAN' prefix properly
    search_string = "SAN" + san_number if not san_number.startswith("SAN") else san_number
    unique = all(search_string != row[0] for row in all_sans_sheet.iter_rows(min_row=2, values_only=True))
    print(f"Checking SAN {search_string}: Unique - {unique}")  # Debug print
    return unique


def show_san_input():
    dialog = SANInputDialog(root, "Enter SAN Number")
    return dialog.result


def open_spreadsheet():
    try:
        if os.name == 'nt':
            os.startfile(workbook_path)
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.run([opener, workbook_path])
    except Exception as e:
        tk.messagebox.showerror("Error", f"Failed to open the spreadsheet: {e}")

frame = ctk.CTkFrame(root)
frame.pack(padx=3, pady=3, fill='both', expand=True)
entry_frame = ctk.CTkFrame(frame)
entry_frame.pack(pady=3)

plots_menu.add_command(label="Open Spreadsheet", command=open_spreadsheet)

button_width = 25
button_1 = ctk.CTkButton(entry_frame, text="Basement 4.2", command=lambda: switch_sheets('original'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_1.pack(side='left', padx=3)
button_2 = ctk.CTkButton(entry_frame, text="Build Room", command=lambda: switch_sheets('backup'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_2.pack(side='left', padx=(3, 50))
button_subtract = ctk.CTkButton(entry_frame, text="-", command=lambda: update_count('subtract'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_subtract.pack(side='left', padx=3)
entry_value = tk.Entry(entry_frame, font=("Helvetica", 14), justify='center', width=5, validate="key", validatecommand=vcmd)
entry_value.pack(side='left', padx=3)
button_add = ctk.CTkButton(entry_frame, text="+", command=lambda: update_count('add'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_add.pack(side='left', padx=3)
# xlsx_button = ctk.CTkButton(entry_frame, text=".xlsx", command=open_spreadsheet, width=button_width, font=("Helvetica", 14))
# xlsx_button.pack(side='left', padx=3)


def update_treeview():
    tree.delete(*tree.get_children())
    workbook = load_workbook(workbook_path)
    item_sheet = workbook[current_sheets[0]]
    row_count = 0
    for row in item_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            if row_count % 2 == 0:
                bg_color = 'white'
            else:
                bg_color = '#f0f0f0'
            tree.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
            tree.tag_configure('oddrow', background='#f0f0f0')
            tree.tag_configure('evenrow', background='white')
            row_count += 1


def log_change(item, action, count=1, san_number="", timestamp_sheet=None):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    action_text = f"{action} {count}" if san_number == "" else f"{action} 1"
    try:
        if timestamp_sheet is not None:
            timestamp_sheet.append([timestamp, item, action_text, san_number])
            workbook.save(workbook_path)
            update_log_view()
            logging.info(f"Logged change: Time: {timestamp}, Item: {item}, Action: {action_text}, SAN: {san_number}")
        else:
            logging.error("No timestamp sheet provided for logging.")
    except Exception as e:
        logging.error(f"Failed to log change: {e}")
        tk.messagebox.showerror("Error", f"Failed to log change: {e}")


def switch_sheets(sheet_type):
    global current_sheets
    current_sheets = sheets[sheet_type]
    update_treeview()
    update_log_view()


def update_log_view():
    if 'log_view' in globals():
        log_view.delete(*log_view.get_children())
        log_sheet = workbook[current_sheets[1]]
        all_rows = list(log_sheet.iter_rows(min_row=2, values_only=True))
        # Adjust the sorting to use the first column (timestamp)
        sorted_rows = sorted(all_rows, key=lambda r: datetime.strptime(r[0], "%Y-%m-%d %H:%M:%S") if r[0] else datetime.min, reverse=True)
        row_count = 0
        for row in sorted_rows:
            if row[0] is not None:
                log_view.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
                log_view.tag_configure('oddrow', background='#f0f0f0')
                log_view.tag_configure('evenrow', background='white')
                row_count += 1


# Validation command for alphanumeric input
alphanumeric_vcmd = (root.register(lambda P: re.match('^[a-zA-Z0-9]*$', P) is not None), '%P')

# Add the Serial Number Input Dialog Class
class SerialNumberInputDialog(tk.Toplevel):
    def __init__(self, parent, title=None):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.parent = parent
        self.result = None
        # Use alphanumeric validation command for entry widget
        self.entry = ttk.Entry(self, validate="key", validatecommand=alphanumeric_vcmd)
        self.create_widgets()
        self.grab_set()
        self.geometry(f"+{parent.winfo_rootx() + parent.winfo_width() // 2 - 100}+{parent.winfo_rooty() + parent.winfo_height() // 2 - 50}")
        self.wait_window(self)

    def create_widgets(self):
        # No longer need to redefine self.entry here as it's already defined in __init__
        self.entry.pack(padx=5, pady=5)
        button_frame = tk.Frame(self)
        button_frame.pack(pady=5)
        submit_button = ttk.Button(button_frame, text="Submit", command=self.on_submit)
        submit_button.pack(side='left', padx=5)
        cancel_button = ttk.Button(button_frame, text="Cancel", command=self.on_cancel)
        cancel_button.pack(side='left', padx=5)

    def on_submit(self):
        serial_number = self.entry.get()
        if serial_number:  # Add more validation if needed
            self.result = serial_number
            self.destroy()
        else:
            tk.messagebox.showerror("Error", "Please enter a valid Serial Number.", parent=self)
            self.entry.focus_set()

    def on_cancel(self):
        self.result = None
        self.destroy()

def show_serial_number_input():
    dialog = SerialNumberInputDialog(root, "Serial #")
    return dialog.result

def save_serial_number(serial_number):
    if 'Headsets' not in workbook.sheetnames:
        workbook.create_sheet('Headsets')
        workbook['Headsets'].append(["Serial Number"])
    workbook['Headsets'].append([serial_number])
    workbook.save(workbook_path)

class NotesInputDialog(tk.Toplevel):
    def __init__(self, parent, title=None):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.parent = parent
        self.result = None
        self.create_widgets()
        self.grab_set()
        self.geometry(f"+{parent.winfo_rootx() + parent.winfo_width() // 2 - 100}+{parent.winfo_rooty() + parent.winfo_height() // 2 - 50}")
        self.wait_window(self)

    def create_widgets(self):
        self.text = tk.Text(self, height=5, width=30)
        self.text.pack(padx=5, pady=5)
        button_frame = tk.Frame(self)
        button_frame.pack(pady=5)
        submit_button = ttk.Button(button_frame, text="Submit", command=self.on_submit)
        submit_button.pack(side='left', padx=5)
        cancel_button = ttk.Button(button_frame, text="Cancel", command=self.on_cancel)
        cancel_button.pack(side='left', padx=5)

    def on_submit(self):
        notes = self.text.get("1.0", tk.END).strip()
        if notes:
            self.result = notes
            self.destroy()
        else:
            tk.messagebox.showerror("Error", "Please enter some notes.", parent=self)

    def on_cancel(self):
        self.result = None
        self.destroy()

def show_notes_input():
    dialog = NotesInputDialog(root, "Notes")
    return dialog.result


class ServiceNowInputDialog(tk.Toplevel):
    def __init__(self, parent, title=None):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.parent = parent
        self.result = None
        self.create_widgets()
        self.grab_set()
        self.geometry(f"+{parent.winfo_rootx() + parent.winfo_width() // 2 - 100}+{parent.winfo_rooty() + parent.winfo_height() // 2 - 50}")
        self.wait_window(self)

    def create_widgets(self):
        ttk.Label(self, text="Ticket Prefix:").pack(padx=5, pady=5)
        self.prefix_var = tk.StringVar()
        self.prefix_menu = ttk.Combobox(self, textvariable=self.prefix_var, values=["TASK", "RITM", "CHG", "INC"])
        self.prefix_menu.pack(padx=5, pady=5)
        self.prefix_menu.current(0)  # Default selection
        ttk.Label(self, text="Ticket Number:").pack(padx=5, pady=5)
        self.entry = ttk.Entry(self)
        self.entry.pack(padx=5, pady=5)
        button_frame = tk.Frame(self)
        button_frame.pack(pady=5)
        submit_button = ttk.Button(button_frame, text="Submit", command=self.on_submit)
        submit_button.pack(side='left', padx=5)
        cancel_button = ttk.Button(button_frame, text="Cancel", command=self.on_cancel)
        cancel_button.pack(side='left', padx=5)

    def on_submit(self):
        ticket_number = self.entry.get()
        prefix = self.prefix_var.get()
        if ticket_number and prefix:
            # Concatenate prefix and ticket number without a dash
            self.result = f"{prefix}{ticket_number}"
            self.destroy()
        else:
            tk.messagebox.showerror("Error", "Please enter a valid Ticket Number.", parent=self)

    def on_cancel(self):
        self.result = None
        self.destroy()

def show_servicenow_input():
    dialog = ServiceNowInputDialog(root, "Enter ServiceNow Ticket")
    return dialog.result

def save_servicenow_ticket(serial_number, servicenow_ticket):
    if 'Headsets' not in workbook.sheetnames:
        workbook.create_sheet('Headsets')
        workbook['Headsets'].append(["Serial Number", "ServiceNow #"])
    headset_sheet = workbook['Headsets']
    # Remove dash from the ServiceNow ticket number before saving
    servicenow_ticket = servicenow_ticket.replace('-', '')
    headset_sheet.append([serial_number, servicenow_ticket])
    workbook.save(workbook_path)

def update_count(operation):
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if selected_item:
        input_value = entry_value.get()
        if input_value.isdigit():
            input_value = int(input_value)
            item_sheet = workbook[current_sheets[0]]
            timestamp_sheet = workbook[current_sheets[1]]

            # Define criteria for when a SAN number is required
            san_required = any(g in selected_item for g in ["G8", "G9", "G10"])
            headset_required = "headset" in selected_item.lower()

            # Logic for headsets
            if headset_required and operation == 'subtract':
                headset_count = 0
                while headset_count < input_value:
                    serial_number = show_serial_number_input()
                    if serial_number is None:  # User cancelled the input
                        return
                    servicenow_ticket = show_servicenow_input()
                    notes = ""
                    if not servicenow_ticket:  # If no ServiceNow ticket is provided
                        notes = show_notes_input()
                        if notes is None:  # If no notes are provided, and the dialog is cancelled
                            return
                    save_headset_info(serial_number, servicenow_ticket, notes)
                    headset_count += 1

            # Logic for SAN required items
            if san_required:
                san_count = 0
                while san_count < input_value:
                    san_number = show_san_input()
                    if san_number is None:  # User cancelled the input
                        return
                    san_number = "SAN" + san_number if not san_number.startswith("SAN") else san_number
 


                    if operation == 'add':
                        if is_san_unique(san_number):
                            print(f"Adding unique SAN {san_number}")  # Debug print
                            all_sans_sheet.append([san_number, selected_item, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                            log_change(selected_item, operation, 1, san_number, timestamp_sheet)  # Correctly pass the SAN number
                            san_count += 1
                        else:
                            tk.messagebox.showerror("Error", f"Duplicate or already used SAN number: {san_number}", parent=root)
                    elif operation == 'subtract':
                        # Check if the SAN number exists in the sheet
                        row_to_delete = None
                        for row in all_sans_sheet.iter_rows(min_row=2):
                            if row[0].value == san_number:
                                row_to_delete = row[0].row
                                break


                        if row_to_delete:
                            all_sans_sheet.delete_rows(row_to_delete)
                            log_change(selected_item, operation, 1, san_number, timestamp_sheet)
                            san_count += 1
                        else:
                            tk.messagebox.showerror("Error", f"That SAN number does not exist in the All SANs sheet: {san_number}", parent=root)
            
            if headset_required and operation == 'add':
                headset_count = 0
                while headset_count < input_value:
                    serial_number = show_serial_number_input()
                    if serial_number is None:  # User cancelled the input
                        return
                    servicenow_ticket = show_servicenow_input()  # Prompt for ServiceNow Ticket
                    if servicenow_ticket is None:  # User cancelled the input
                        return
                    save_serial_number(serial_number)
                    save_servicenow_ticket(serial_number, servicenow_ticket)
                    headset_count += 1
                    
            # Adjust item counts for non-SAN or non-headset items
            if not san_required and not headset_required:
                for row in item_sheet.iter_rows(min_row=2):
                    if row[0].value == selected_item:
                        row[1].value = row[2].value or 0
                        if operation == 'add':
                            row[2].value = (row[2].value or 0) + input_value
                        elif operation == 'subtract':
                            row[2].value = max((row[2].value or 0) - input_value, 0)
                        log_change(selected_item, operation, input_value, "", timestamp_sheet)

            workbook.save(workbook_path)
            update_treeview()
            update_log_view()    

            # Log the change for items not requiring SAN
            if not san_required:
                log_change(selected_item, operation, input_value, "", timestamp_sheet)

            workbook.save(workbook_path)
            update_treeview()
            update_log_view()

def save_headset_info(serial_number, servicenow_ticket, notes):
    if 'Headsets' not in workbook.sheetnames:
        workbook.create_sheet('Headsets')
        workbook['Headsets'].append(["Serial Number", "ServiceNow #", "Notes"])
    headset_sheet = workbook['Headsets']
    headset_sheet.append([serial_number, servicenow_ticket, notes])
    workbook.save(workbook_path)

columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', style="Treeview")
for col in columns:
    tree.heading(col, text=col, anchor='w')
    tree.column("Item", anchor='w', width=250, stretch=False) # Width of the "Item" column in the treeview. The other columns are default width.
    tree.column("LastCount", anchor='w', width=175, stretch=False)
tree.pack(expand=True, fill="both", padx=3, pady=3)

log_view_frame = ctk.CTkFrame(root)
log_view_frame.pack(side=tk.BOTTOM, fill='both', expand=True, padx=10, pady=10)

log_view_columns = ("Timestamp", "Item", "Action", "SAN Number")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", style="Treeview", height=8)
for col in log_view_columns:
    log_view.heading(col, text=col, anchor='w')
    log_view.column("Timestamp", anchor='w', width=190, stretch=False)
    log_view.column("Item", anchor='w', width=160, stretch=False)
    log_view.column("Action", anchor='w', width=100, stretch=False)

scrollbar_log = ttk.Scrollbar(log_view_frame, orient="vertical", command=log_view.yview)
scrollbar_log.pack(side='right', fill='y')
log_view.configure(yscrollcommand=scrollbar_log.set)
log_view.pack(expand=True, fill='both')

root.after(100, update_treeview)
update_log_view()

root.mainloop()

