# Build Room\build_roomv2.18.py
# Author: Macdara O Murchu
# 20.02.24

import logging.config
from pathlib import Path
from tkinter import Menu
import customtkinter as ctk
import os
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from datetime import datetime
import subprocess
import tkinter.simpledialog as sd

logging_conf_path = Path('logging.conf')
if logging_conf_path.exists() and logging_conf_path.stat().st_size > 0:
    try:
        logging.config.fileConfig(logging_conf_path)
    except Exception as e:
        logging.error(f"Error configuring logging: {e}", exc_info=True)
else:
    logging.basicConfig(level=logging.DEBUG)

def run_inventory_script():
    script_path = script_directory / "inventory-levels_4.2v1.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_4.2v1.py' does not exist in the directory.")

def run_build_room_inventory_script():
    script_path = script_directory / "inventory-levels_BRv1.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_BRv1.py' does not exist in the directory.")

def run_combined_rooms_inventory_script():
    script_path = script_directory / "inventory-levels_combinedv1.1.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_combinedv1.py' does not exist in the directory.")

def view_headsets_log():
    log_window = tk.Toplevel(root)
    log_window.title("Headsets In Stock")
    log_window.geometry("625x750")

    # Create a Treeview widget to display the log
    columns = ("Serial #", "ServiceNow #", "Notes")
    log_tree = ttk.Treeview(log_window, columns=columns, show="headings")
    for col in columns:
        log_tree.heading(col, text=col)
        log_tree.column(col, anchor="center")
    log_tree.pack(expand=True, fill="both", padx=10, pady=10)

    # Scrollbar for the Treeview
    scrollbar = ttk.Scrollbar(log_window, orient="vertical", command=log_tree.yview)
    scrollbar.pack(side="right", fill="y")
    log_tree.configure(yscrollcommand=scrollbar.set)

    # Load and display data from the "Headsets" sheet
    if 'Headsets' in workbook.sheetnames:
        headsets_sheet = workbook['Headsets']
        for row in headsets_sheet.iter_rows(min_row=2, values_only=True):
            log_tree.insert('', 'end', values=row)
    else:
        tk.messagebox.showinfo("Info", "Headsets log is empty.", parent=log_window)

def view_all_sans_log():
    log_window = tk.Toplevel(root)
    log_window.title("SANs In Stock")
    log_window.geometry("625x750")

    # Create a Treeview widget to display the log
    columns = ("SAN Number", "Item", "Timestamp")
    log_tree = ttk.Treeview(log_window, columns=columns, show="headings")
    for col in columns:
        log_tree.heading(col, text=col)
        log_tree.column(col, anchor="center")
    log_tree.pack(expand=True, fill="both", padx=10, pady=10)

    # Scrollbar for the Treeview
    scrollbar = ttk.Scrollbar(log_window, orient="vertical", command=log_tree.yview)
    scrollbar.pack(side="right", fill="y")
    log_tree.configure(yscrollcommand=scrollbar.set)

    # Load and display data from the "All SANs" sheet
    if 'All SANs' in workbook.sheetnames:
        all_sans_sheet = workbook['All SANs']
        for row in all_sans_sheet.iter_rows(min_row=2, values_only=True):
            log_tree.insert('', 'end', values=row)
    else:
        tk.messagebox.showinfo("Info", "All SANs log is empty.", parent=log_window)

root = ctk.CTk()
root.title("Perth EUC Assets")
root.geometry("750x750")

menu_bar = tk.Menu(root)
plots_menu = tk.Menu(menu_bar, tearoff=0)
plots_menu.add_command(label="Basement 4.2 Inventory", command=run_inventory_script)
plots_menu.add_command(label="Build Room Inventory", command=run_build_room_inventory_script)
plots_menu.add_command(label="Combined Inventory", command=run_combined_rooms_inventory_script)
plots_menu.add_command(label="SANs In Stock", command=view_all_sans_log)
plots_menu.add_command(label="Headsets In Stock", command=view_headsets_log)
menu_bar.add_cascade(label="Data", menu=plots_menu)
root.config(menu=menu_bar)

script_directory = Path(__file__).parent
workbook_path = script_directory / 'EUC_Perth_Assets.xlsx'
if Path(workbook_path).exists():
    workbook = load_workbook(workbook_path)
else:
    workbook = Workbook()
    workbook.active.title = '4.2 Items'
    workbook.create_sheet('4.2 Timestamps')
    workbook.create_sheet('BR Items')
    workbook.create_sheet('BR Timestamps')
    workbook.create_sheet('Project Designated Items')
    workbook.create_sheet('Project Designated Timestamps')
    workbook.create_sheet('All SANs')
    workbook['4.2 Items'].append(["Item", "LastCount", "NewCount"])
    workbook['4.2 Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['BR Items'].append(["Item", "LastCount", "NewCount"])
    workbook['BR Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['Project Designated Items'].append(["Item", "LastCount", "NewCount"])
    workbook['Project Designated Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['All SANs'].append(["SAN Number", "Item", "Timestamp"])
    workbook.save(workbook_path)

all_sans_sheet = workbook['All SANs']
sheets = {'original': ('4.2 Items', '4.2 Timestamps'), 'backup': ('BR Items', 'BR Timestamps')}
current_sheets = sheets['original']

style = ttk.Style()
style.configure("Treeview", font=('Helvetica', 12))

vcmd = (root.register(lambda P: P.isdigit() or P == ""), '%P')

class SANInputDialog(tk.Toplevel):
    def __init__(self, parent, title=None):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.parent = parent
        self.result = None
        self.create_widgets()
        self.grab_set()
        self.geometry(f"+{parent.winfo_rootx() + parent.winfo_width() // 2 - 100}+{parent.winfo_rooty() + parent.winfo_height() // 2 - 50}")
        self.wait_window(self)

    def create_widgets(self):
        self.entry = ttk.Entry(self, validate="key", validatecommand=vcmd)
        self.entry.pack(padx=5, pady=5)
        button_frame = tk.Frame(self)
        button_frame.pack(pady=5)
        submit_button = ttk.Button(button_frame, text="Submit", command=self.on_submit)
        submit_button.pack(side='left', padx=5)
        cancel_button = ttk.Button(button_frame, text="Cancel", command=self.on_cancel)
        cancel_button.pack(side='left', padx=5)

    def on_submit(self):
        san_input = self.entry.get()
        if san_input and len(san_input) >= 5 and len(san_input) <= 6:
            self.result = san_input
            self.destroy()
        else:
            tk.messagebox.showerror("Error", "Please enter a valid SAN number.", parent=self)
            self.entry.focus_set()

    def on_cancel(self):
        self.result = None
        self.destroy()

class ServiceNowInputDialog(tk.Toplevel):
    def __init__(self, parent, title=None):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.parent = parent
        self.result = None
        self.create_widgets()
        self.grab_set()
        self.geometry(f"+{parent.winfo_rootx() + parent.winfo_width() // 2 - 100}+{parent.winfo_rooty() + parent.winfo_height() // 2 - 50}")
        self.wait_window(self)

    def create_widgets(self):
        # Dropdown menu for prefix selection
        self.prefix_var = tk.StringVar(self, value="TASK")  # Set default value to "TASK"
        self.prefix_dropdown = tk.OptionMenu(self, self.prefix_var, "TASK", "RITM")
        self.prefix_dropdown.pack(padx=10, pady=5)

        # Entry for ServiceNow number
        self.number_entry = tk.Entry(self)
        self.number_entry.pack(padx=10, pady=5)

        # Submit button
        submit_button = tk.Button(self, text="Submit", command=self.on_submit)
        submit_button.pack(pady=5)

    def on_submit(self):
        prefix = self.prefix_var.get()
        number = self.number_entry.get()
        self.result = f"{prefix} {number}"
        self.destroy()

def is_san_unique(san_number):
    # Adjust the search to account for the 'SAN' prefix properly
    search_string = "SAN" + san_number if not san_number.startswith("SAN") else san_number
    unique = all(search_string != row[0] for row in all_sans_sheet.iter_rows(min_row=2, values_only=True))
    print(f"Checking SAN {search_string}: Unique - {unique}")  # Debug print
    return unique

def show_san_input():
    dialog = SANInputDialog(root, "SAN #")
    return dialog.result

def open_spreadsheet():
    try:
        if os.name == 'nt':
            os.startfile(workbook_path)
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.run([opener, workbook_path])
    except Exception as e:
        tk.messagebox.showerror("Error", f"Failed to open the spreadsheet: {e}")

frame = ctk.CTkFrame(root)
frame.pack(padx=3, pady=3, fill='both', expand=True)
entry_frame = ctk.CTkFrame(frame)
entry_frame.pack(pady=3)

plots_menu.add_command(label="Open Spreadsheet", command=open_spreadsheet)

button_width = 25
button_1 = ctk.CTkButton(entry_frame, text="Basement 4.2", command=lambda: switch_sheets('original'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_1.pack(side='left', padx=3)
button_2 = ctk.CTkButton(entry_frame, text="Build Room", command=lambda: switch_sheets('backup'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_2.pack(side='left', padx=(3, 50))
button_subtract = ctk.CTkButton(entry_frame, text="-", command=lambda: update_count('subtract'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_subtract.pack(side='left', padx=3)
entry_value = tk.Entry(entry_frame, font=("Helvetica", 14), justify='center', width=5, validate="key", validatecommand=vcmd)
entry_value.pack(side='left', padx=3)
button_add = ctk.CTkButton(entry_frame, text="+", command=lambda: update_count('add'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_add.pack(side='left', padx=3)

def update_treeview():
    tree.delete(*tree.get_children())
    workbook = load_workbook(workbook_path)
    item_sheet = workbook[current_sheets[0]]
    row_count = 0
    for row in item_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            if row_count % 2 == 0:
                bg_color = 'white'
            else:
                bg_color = '#f0f0f0'
            tree.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
            tree.tag_configure('oddrow', background='#f0f0f0')
            tree.tag_configure('evenrow', background='white')
            row_count += 1

def log_change(item, action, count=1, san_numbers=[], serial_numbers=[], servicenow_numbers=[], timestamp_sheet=None):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Handling operations logging for SAN Numbers
    for san_number in san_numbers:
        action_text = f"{action} 1 SAN Device"
        try:
            timestamp_sheet.append([timestamp, item, action_text, san_number, "", servicenow_numbers[0] if servicenow_numbers else ""])
        except Exception as e:
            logging.error(f"Failed to log SAN device operation: {e}")
    
    # Handling operations logging for Serial Numbers (Headsets)
    for serial_number in serial_numbers:
        action_text = f"{action} 1 Headset"
        try:
            timestamp_sheet.append([timestamp, item, action_text, "", serial_number, servicenow_numbers[0] if servicenow_numbers else ""])
        except Exception as e:
            logging.error(f"Failed to log Headset operation: {e}")
    
    # Handling operations for items without specific identifiers (Single ServiceNow prompt for multiple items if needed)
    if not san_numbers and not serial_numbers:
        action_text = f"{action} {count}"
        try:
            timestamp_sheet.append([timestamp, item, action_text, "", "", servicenow_numbers[0] if servicenow_numbers else ""])
        except Exception as e:
            logging.error(f"Failed to log operation: {e}")
            
    workbook.save(workbook_path)
    update_log_view()
    logging.info(f"Logged change: Time: {timestamp}, Item: {item}, Action: {action}")
def switch_sheets(sheet_type):
    global current_sheets
    current_sheets = sheets[sheet_type]
    update_treeview()
    update_log_view()

def update_log_view():
    if 'log_view' in globals():
        log_view.delete(*log_view.get_children())
        log_sheet = workbook[current_sheets[1]]
        all_rows = list(log_sheet.iter_rows(min_row=2, values_only=True))
        sorted_rows = sorted(
            all_rows, 
            key=lambda r: datetime.strptime(r[0], "%Y-%m-%d %H:%M:%S") if r[0] else datetime.min, 
            reverse=True
        )
        
        row_count = 0  # Initialize row_count here
        for row in sorted_rows:
            if row[0] is not None:
                log_view.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
                log_view.tag_configure('oddrow', background='#f0f0f0')
                log_view.tag_configure('evenrow', background='white')
                row_count += 1  # Increment row_count for each row processed

# Serial Number Input Dialog Function
# Function to prompt for a Serial Number
def serial_number_input():
    serial_number = sd.askstring("Serial Number", "Enter Serial Number:", parent=root)
    # Validate the serial number
    if serial_number and len(serial_number) == 6 and serial_number.isalnum():
        return serial_number  # Return the valid Serial Number
    else:
        tk.messagebox.showerror("Error", "Invalid Serial Number. Please enter 6 alphanumeric characters.")
        return None  # Return None if invalid or no input

# Function to prompt for a ServiceNow Number
def servicenow_number_input():
    servicenow_number = sd.askstring("ServiceNow Number", "Enter ServiceNow Number:", parent=root)
    # No extensive validation here for simplicity, customization is possible based on given requirements or format
    if servicenow_number:
        return servicenow_number  # Return the ServiceNow Number
    else:
        return None  # Return None if no input

        # Dropdown menu for prefix selection
        self.prefix_var = tk.StringVar()
        self.prefix_dropdown = ttk.Combobox(self, textvariable=self.prefix_var, state="readonly", values=["TASK", "RITM"])
        self.prefix_dropdown.grid(row=0, column=1, padx=10, pady=5)
        self.prefix_dropdown.set("TASK")  # default value

        # Entry for ServiceNow number
        self.number_entry = ttk.Entry(self)
        self.number_entry.grid(row=1, column=1, padx=10, pady=5)

        # Submit button
        submit_button = ttk.Button(self, text="Submit", command=self.on_submit)
        submit_button.grid(row=2, column=1, padx=10, pady=5)

    def on_submit(self):
        prefix = self.prefix_var.get()
        number = self.number_entry.get()
        if 6 <= len(number) <= 8 and number.isdigit():
            self.result = f"{prefix}{number}"
            self.destroy()
        else:
            tk.messagebox.showerror("Error", "Please enter a 7-digit number.", parent=self)

    def show(self):
        self.wm_deiconify()
        self.number_entry.focus_force()
        self.wait_window()
        return self.result
    
def servicenow_number_input():
    dialog = ServiceNowInputDialog(root)
    return dialog.show()

def update_count(operation):
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if selected_item:
        san_numbers, serial_numbers, servicenow_numbers = [], [], []
        input_value = entry_value.get()
        if input_value.isdigit():
            input_value = int(input_value)
            # Prompting for SAN, Serial, ServiceNow numbers
            if "Headset" in selected_item:
                for _ in range(input_value):
                    serial_number = serial_number_input()
                    if serial_number is None:  # User cancelled the operation
                        return
                    serial_numbers.append(serial_number)
                servicenow_number = servicenow_number_input()
                if servicenow_number is not None:
                    servicenow_numbers.append(servicenow_number)
            elif any(generation in selected_item for generation in ["G8", "G9", "G10"]):
                for _ in range(input_value):
                    san_number = show_san_input()
                    if san_number is None:  # User cancelled the operation
                        return
                    san_numbers.append(san_number)
                servicenow_number = servicenow_number_input()
                if servicenow_number is not None:
                    servicenow_numbers.append(servicenow_number)
            else:
                # Non-SAN and Non-Headset items, ServiceNow number is prompted once
                servicenow_number = servicenow_number_input()
                if servicenow_number is not None:
                    servicenow_numbers.append(servicenow_number)
                    
            # Update item count and logging the operation
            item_sheet = workbook[current_sheets[0]]
            timestamp_sheet = workbook[current_sheets[1]]
            for row in item_sheet.iter_rows(min_row=2):
                if row[0].value == selected_item:
                    row[1].value = row[2].value or 0  # Update LastCount
                    # Update NewCount based on the operation
                    if operation == 'add':
                        row[2].value = (row[2].value or 0) + input_value
                    elif operation == 'subtract':
                        row[2].value = max((row[2].value or 0) - input_value, 0)
                    workbook.save(workbook_path)
                   
            # Log the operations
            log_change(selected_item, operation,
                       count=input_value,
                       san_numbers=san_numbers,
                       serial_numbers=serial_numbers,
                       servicenow_numbers=servicenow_numbers,
                       timestamp_sheet=timestamp_sheet)
                    
            update_treeview()
            update_log_view()
        else:
            tk.messagebox.showerror("Invalid Input", "Please enter a numeric value for the count.")


columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', style="Treeview")
for col in columns:
    tree.heading(col, text=col, anchor='center')
    tree.column("Item", anchor="center", width=250, stretch=False)
    tree.column("LastCount", anchor="center", width=175, stretch=False)
    tree.column("NewCount", anchor="center", width=175, stretch=False)  # Ensure this line is correctly specifying "center"
tree.pack(expand=True, fill="both", padx=3, pady=3)

log_view_frame = ctk.CTkFrame(root)
log_view_frame.pack(side=tk.BOTTOM, fill='both', expand=True, padx=10, pady=10)

log_view_columns = ("Timestamp", "Item", "Action", "SAN Number", "Serial #", "ServiceNow #")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", style="Treeview", height=12)
for col in log_view_columns:
    log_view.heading(col, text=col, anchor='center')
# Now adjust width as needed after log_view has been instantiated
# log_view.column("Timestamp", anchor='center', width=175)
# log_view.column("Item", anchor='center', width=130)
# log_view.column("Action", anchor='center', width=50)
# log_view.column("SAN Number", anchor='center', width=70)
# log_view.column("Serial #", anchor='center', width=70)
# log_view.column("ServiceNow #", anchor='center', width=95)
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", style="Treeview", height=12)
# for col in log_view_columns:
#     log_view.heading(col, text=col, anchor='center')
#     # Adjust width as needed
#     log_view.column(col, anchor='center', width=11)

scrollbar_log = ttk.Scrollbar(log_view_frame, orient="vertical", command=log_view.yview)
scrollbar_log.pack(side='right', fill='y')
log_view.configure(yscrollcommand=scrollbar_log.set)
log_view.pack(expand=True, fill='both')

root.after(100, update_treeview)
update_log_view()

root.mainloop()