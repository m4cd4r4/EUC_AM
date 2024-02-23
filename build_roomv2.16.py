# Add logic for ServiceNow dialog for all items. If servicenow cancelled, new dialog for notes. If notes cancelled, none.

# Build Room\build_roomv2.16.py
# Author: Macdara O Murchu
# 18.02.24

import logging.config
from pathlib import Path
from tkinter import Menu
import customtkinter as ctk
import os
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from datetime import datetime
import subprocess
import tkinter.simpledialog as sd

logging_conf_path = Path('logging.conf')
if logging_conf_path.exists() and logging_conf_path.stat().st_size > 0:
    try:
        logging.config.fileConfig(logging_conf_path)
    except Exception as e:
        logging.error(f"Error configuring logging: {e}", exc_info=True)
else:
    logging.basicConfig(level=logging.DEBUG)

def run_inventory_script():
    script_path = script_directory / "inventory-levels_4.2v1.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_4.2v1.py' does not exist in the directory.")

def run_build_room_inventory_script():
    script_path = script_directory / "inventory-levels_BRv1.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_BRv1.py' does not exist in the directory.")

def run_combined_rooms_inventory_script():
    script_path = script_directory / "inventory-levels_combinedv1.1.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_combinedv1.py' does not exist in the directory.")

def view_headsets_log():
    log_window = tk.Toplevel(root)
    log_window.title("Headsets In Stock")
    log_window.geometry("600x400")

    # Update the columns to include "Notes"
    columns = ("Serial #", "ServiceNow #", "Notes")
    log_tree = ttk.Treeview(log_window, columns=columns, show="headings")
    for col in columns:
        log_tree.heading(col, text=col)
        log_tree.column(col, anchor="w")
    log_tree.pack(expand=True, fill="both", padx=10, pady=10)

    # Scrollbar for the Treeview
    scrollbar = ttk.Scrollbar(log_window, orient="vertical", command=log_tree.yview)
    scrollbar.pack(side="right", fill="y")
    log_tree.configure(yscrollcommand=scrollbar.set)

    # Load and display data from the "Headsets" sheet, now including the "Notes" column
    if 'Headsets' in workbook.sheetnames:
        headsets_sheet = workbook['Headsets']
        for row in headsets_sheet.iter_rows(min_row=2, values_only=True):
            # Ensure the row has enough values to include the "Notes" column; pad with empty strings if necessary
            padded_row = row + ('',) * (len(columns) - len(row))
            log_tree.insert('', 'end', values=padded_row)
    else:
        tk.messagebox.showinfo("Info", "Headsets log is empty.", parent=log_window)


def view_all_sans_log():
    log_window = tk.Toplevel(root)
    log_window.title("SANs In Stock")
    log_window.geometry("600x400")

    # Create a Treeview widget to display the log
    columns = ("SAN Number", "Item", "Timestamp")
    log_tree = ttk.Treeview(log_window, columns=columns, show="headings")
    for col in columns:
        log_tree.heading(col, text=col)
        log_tree.column(col, anchor="w")
    log_tree.pack(expand=True, fill="both", padx=10, pady=10)

    # Scrollbar for the Treeview
    scrollbar = ttk.Scrollbar(log_window, orient="vertical", command=log_tree.yview)
    scrollbar.pack(side="right", fill="y")
    log_tree.configure(yscrollcommand=scrollbar.set)

    # Load and display data from the "All SANs" sheet
    if 'All SANs' in workbook.sheetnames:
        all_sans_sheet = workbook['All SANs']
        for row in all_sans_sheet.iter_rows(min_row=2, values_only=True):
            log_tree.insert('', 'end', values=row)
    else:
        tk.messagebox.showinfo("Info", "All SANs log is empty.", parent=log_window)

root = ctk.CTk()
root.title("Perth EUC Assets")
root.geometry("500x650")

menu_bar = tk.Menu(root)
plots_menu = tk.Menu(menu_bar, tearoff=0)
plots_menu.add_command(label="Basement 4.2 Inventory", command=run_inventory_script)
plots_menu.add_command(label="Build Room Inventory", command=run_build_room_inventory_script)
plots_menu.add_command(label="Combined Inventory", command=run_combined_rooms_inventory_script)
plots_menu.add_command(label="SANs In Stock", command=view_all_sans_log)
plots_menu.add_command(label="Headsets In Stock", command=view_headsets_log)
menu_bar.add_cascade(label="Data", menu=plots_menu)
root.config(menu=menu_bar)

script_directory = Path(__file__).parent
workbook_path = script_directory / 'EUC_Perth_Assets.xlsx'
if Path(workbook_path).exists():
    workbook = load_workbook(workbook_path)
else:
    workbook = Workbook()
    workbook.active.title = '4.2 Items'
    workbook.create_sheet('4.2 Timestamps')
    workbook.create_sheet('BR Items')
    workbook.create_sheet('BR Timestamps')
    workbook.create_sheet('Project Designated Items')
    workbook.create_sheet('Project Designated Timestamps')
    workbook.create_sheet('All SANs')
    workbook['4.2 Items'].append(["Item", "LastCount", "NewCount"])
    workbook['4.2 Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['BR Items'].append(["Item", "LastCount", "NewCount"])
    workbook['BR Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['Project Designated Items'].append(["Item", "LastCount", "NewCount"])
    workbook['Project Designated Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['All SANs'].append(["SAN Number", "Item", "Timestamp"])
    workbook.save(workbook_path)

all_sans_sheet = workbook['All SANs']
sheets = {'original': ('4.2 Items', '4.2 Timestamps'), 'backup': ('BR Items', 'BR Timestamps')}
current_sheets = sheets['original']

style = ttk.Style()
style.configure("Treeview", font=('Helvetica', 12))

vcmd = (root.register(lambda P: P.isdigit() or P == ""), '%P')

class SANInputDialog(tk.Toplevel):
    def __init__(self, parent, title=None):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.parent = parent
        self.result = None
        self.create_widgets()
        self.grab_set()
        self.geometry(f"+{parent.winfo_rootx() + parent.winfo_width() // 2 - 100}+{parent.winfo_rooty() + parent.winfo_height() // 2 - 50}")
        self.wait_window(self)

    def create_widgets(self):
        self.entry = ttk.Entry(self, validate="key", validatecommand=vcmd)
        self.entry.pack(padx=5, pady=5)
        button_frame = tk.Frame(self)
        button_frame.pack(pady=5)
        submit_button = ttk.Button(button_frame, text="Submit", command=self.on_submit)
        submit_button.pack(side='left', padx=5)
        cancel_button = ttk.Button(button_frame, text="Cancel", command=self.on_cancel)
        cancel_button.pack(side='left', padx=5)

    def on_submit(self):
        san_input = self.entry.get()
        if san_input and len(san_input) >= 5 and len(san_input) <= 6:
            self.result = san_input
            self.destroy()
        else:
            tk.messagebox.showerror("Error", "Please enter a valid SAN number.", parent=self)
            self.entry.focus_set()

    def on_cancel(self):
        self.result = None
        self.destroy()

def is_san_unique(san_number):
    # Adjust the search to account for the 'SAN' prefix properly
    search_string = "SAN" + san_number if not san_number.startswith("SAN") else san_number
    unique = all(search_string != row[0] for row in all_sans_sheet.iter_rows(min_row=2, values_only=True))
    print(f"Checking SAN {search_string}: Unique - {unique}")  # Debug print
    return unique

def show_san_input():
    dialog = SANInputDialog(root, "SAN #")
    return dialog.result

def open_spreadsheet():
    try:
        if os.name == 'nt':
            os.startfile(workbook_path)
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.run([opener, workbook_path])
    except Exception as e:
        tk.messagebox.showerror("Error", f"Failed to open the spreadsheet: {e}")

frame = ctk.CTkFrame(root)
frame.pack(padx=3, pady=3, fill='both', expand=True)
entry_frame = ctk.CTkFrame(frame)
entry_frame.pack(pady=3)

plots_menu.add_command(label="Open Spreadsheet", command=open_spreadsheet)

button_width = 25
button_1 = ctk.CTkButton(entry_frame, text="Basement 4.2", command=lambda: switch_sheets('original'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_1.pack(side='left', padx=3)
button_2 = ctk.CTkButton(entry_frame, text="Build Room", command=lambda: switch_sheets('backup'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_2.pack(side='left', padx=(3, 50))
button_subtract = ctk.CTkButton(entry_frame, text="-", command=lambda: update_count('subtract'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_subtract.pack(side='left', padx=3)
entry_value = tk.Entry(entry_frame, font=("Helvetica", 14), justify='center', width=5, validate="key", validatecommand=vcmd)
entry_value.pack(side='left', padx=3)
button_add = ctk.CTkButton(entry_frame, text="+", command=lambda: update_count('add'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_add.pack(side='left', padx=3)

def update_treeview():
    tree.delete(*tree.get_children())
    workbook = load_workbook(workbook_path)
    item_sheet = workbook[current_sheets[0]]
    row_count = 0
    for row in item_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            if row_count % 2 == 0:
                bg_color = 'white'
            else:
                bg_color = '#f0f0f0'
            tree.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
            tree.tag_configure('oddrow', background='#f0f0f0')
            tree.tag_configure('evenrow', background='white')
            row_count += 1

def log_change(item, action, count=1, san_number="", timestamp_sheet=None):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    action_text = f"{action} {count}" if san_number == "" else f"{action} 1"
    try:
        if timestamp_sheet is not None:
            timestamp_sheet.append([timestamp, item, action_text, san_number])
            workbook.save(workbook_path)
            update_log_view()
            logging.info(f"Logged change: Time: {timestamp}, Item: {item}, Action: {action_text}, SAN: {san_number}")
        else:
            logging.error("No timestamp sheet provided for logging.")
    except Exception as e:
        logging.error(f"Failed to log change: {e}")
        tk.messagebox.showerror("Error", f"Failed to log change: {e}")

def switch_sheets(sheet_type):
    global current_sheets
    current_sheets = sheets[sheet_type]
    update_treeview()
    update_log_view()

def update_log_view():
    if 'log_view' in globals():
        log_view.delete(*log_view.get_children())
        log_sheet = workbook[current_sheets[1]]
        all_rows = list(log_sheet.iter_rows(min_row=2, values_only=True))
        # Adjust the sorting to use the first column (timestamp)
        sorted_rows = sorted(all_rows, key=lambda r: datetime.strptime(r[0], "%Y-%m-%d %H:%M:%S") if r[0] else datetime.min, reverse=True)
        row_count = 0
        for row in sorted_rows:
            if row[0] is not None:
                log_view.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
                log_view.tag_configure('oddrow', background='#f0f0f0')
                log_view.tag_configure('evenrow', background='white')
                row_count += 1

# Serial Number Input Dialog Function
def serial_number_input():
    while True:
        serial_num = sd.askstring("Serial Number", "Enter Serial Number:", parent=root)
        if serial_num is None:  # User pressed cancel
            return None
        if len(serial_num) == 6 and serial_num.isalnum():
            return serial_num  # Valid input
        else:
            tk.messagebox.showerror("Invalid Input", "Serial Number must be 6 characters.")

class ServiceNowInputDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.result = None
        self.title("ServiceNow #")

        # Dropdown menu for prefix selection
        self.prefix_var = tk.StringVar()
        self.prefix_dropdown = ttk.Combobox(self, textvariable=self.prefix_var, state="readonly", values=["TASK", "RITM"])
        self.prefix_dropdown.grid(row=0, column=1, padx=10, pady=5)
        self.prefix_dropdown.set("TASK")  # default value

        # Entry for ServiceNow number
        self.number_entry = ttk.Entry(self)
        self.number_entry.grid(row=1, column=1, padx=10, pady=5)

        # Submit button
        submit_button = ttk.Button(self, text="Submit", command=self.on_submit)
        submit_button.grid(row=2, column=1, padx=10, pady=5)

    def on_submit(self):
        prefix = self.prefix_var.get()
        number = self.number_entry.get()
        if 6 <= len(number) <= 8 and number.isdigit():
            self.result = f"{prefix}{number}"
            self.destroy()
        else:
            tk.messagebox.showerror("Error", "Please enter a 7-digit number.", parent=self)

    def show(self):
        self.wm_deiconify()
        self.number_entry.focus_force()
        self.wait_window()
        return self.result
    
def servicenow_number_input():
    dialog = ServiceNowInputDialog(root)
    return dialog.show()

def additional_info_input():
    return tk.simpledialog.askstring("Additional Info", "Enter additional information:", parent=root)

def update_count(operation):
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if not selected_item:
        return

    # Define headsets_sheet within the function scope
    headsets_sheet = workbook['Headsets']

    input_value_str = entry_value.get()
    if input_value_str.isdigit():
        input_value = int(input_value_str)
        item_sheet = workbook[current_sheets[0]]
        timestamp_sheet = workbook[current_sheets[1]]

        # Adjust the logic for handling headsets specifically
        if "Headset" in selected_item:
            for _ in range(input_value):
                serial_number = serial_number_input()
                if not serial_number:  # If serial number input was cancelled
                    continue

                # Adjust for subtract operation
                if operation == 'subtract':
                    servicenow_number = servicenow_number_input()
                    if servicenow_number:
                        new_row = [serial_number, servicenow_number]
                    else:
                        additional_info = additional_info_input()
                        if additional_info:
                            new_row = [serial_number, "", additional_info]
                        else:
                            continue  # Skip if additional info input was cancelled
                    headsets_sheet.append(new_row)
                else:  # For add operation
                    headsets_sheet.append([serial_number])

            # Adjust the item counts for non-headset or if no SAN is required
            if not "Headset" in selected_item or not san_required:
                for row in item_sheet.iter_rows(min_row=2):
                    if row[0].value == selected_item:
                        last_count = row[1].value or 0
                        new_count = (row[2].value or 0) + (input_value if operation == 'add' else -input_value)
                        row[1].value, row[2].value = last_count, max(new_count, 0)  # Prevent negative counts
                        log_change(selected_item, operation, input_value, "", timestamp_sheet)

            workbook.save(workbook_path)
            update_treeview()
            update_log_view()


columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', style="Treeview")
for col in columns:
    tree.heading(col, text=col, anchor='w')
    tree.column("Item", anchor='w', width=250, stretch=False) # Width of the "Item" column in the treeview. The other columns are default width.
    tree.column("LastCount", anchor='w', width=175, stretch=False)
tree.pack(expand=True, fill="both", padx=3, pady=3)

log_view_frame = ctk.CTkFrame(root)
log_view_frame.pack(side=tk.BOTTOM, fill='both', expand=True, padx=10, pady=10)

log_view_columns = ("Timestamp", "Item", "Action", "SAN Number")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", style="Treeview", height=8)
for col in log_view_columns:
    log_view.heading(col, text=col, anchor='w')
    log_view.column("Timestamp", anchor='w', width=190, stretch=False)
    log_view.column("Item", anchor='w', width=160, stretch=False)
    log_view.column("Action", anchor='w', width=100, stretch=False)

scrollbar_log = ttk.Scrollbar(log_view_frame, orient="vertical", command=log_view.yview)
scrollbar_log.pack(side='right', fill='y')
log_view.configure(yscrollcommand=scrollbar_log.set)
log_view.pack(expand=True, fill='both')

root.after(100, update_treeview)
update_log_view()

root.mainloop()