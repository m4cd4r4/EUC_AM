import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from openpyxl import load_workbook

# Load Excel workbook and worksheet
workbook = load_workbook('assets.xlsx')
sheet = workbook.active

# Function to get asset types from the sheet
def get_asset_types():
    return [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]

# Function to add new asset type
def add_new_asset_type():
    new_type = entry_new_asset.get()
    if new_type:
        sheet.append([new_type, 0])  # Add new asset with initial quantity of 0
        workbook.save('assets.xlsx')
        combobox_assets['values'] = get_asset_types()  # Update dropdown values
        entry_new_asset.delete(0, tk.END)
    else:
        messagebox.showwarning("Warning", "Asset type cannot be empty.")

# Placeholder function for 'Add Model' button
def add_model():
    messagebox.showinfo("Info", "Add Model functionality not implemented.")

# Function to update quantity
def update_quantity(addition=True):
    selected_asset = combobox_assets.get()
    if not selected_asset:
        messagebox.showwarning("Warning", "No asset selected.")
        return

    try:
        change = int(entry_quantity.get())
    except ValueError:
        messagebox.showwarning("Warning", "Invalid quantity.")
        return

    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == selected_asset:
            if addition:
                row[1].value += change
            else:
                row[1].value -= change
            workbook.save('assets.xlsx')
            break

# Initialize Tkinter root
root = tk.Tk()
root.title("IT Asset Management")

# Set the initial size of the window (assuming the original size is multiplied by 2)
root.geometry("800x600")  # Adjust this size as needed

# Dropdown for assets
combobox_assets = ttk.Combobox(root, values=get_asset_types(), width=40)
combobox_assets.grid(row=0, columnspan=2, padx=20, pady=20)

# Entry field for new asset type
entry_new_asset = tk.Entry(root, width=40)
entry_new_asset.grid(row=1, columnspan=2, padx=20, pady=10)

# Entry field for quantity
entry_quantity = tk.Entry(root, width=40)
entry_quantity.grid(row=4, columnspan=2, padx=20, pady=10)

# Button for adding new asset type
button_add_asset = tk.Button(root, text="Add Asset Type", command=add_new_asset_type, width=20, height=2)
button_add_asset.grid(row=2, column=0, padx=20, pady=10)

# Button for adding model (next to 'Add Asset Type')
button_add_model = tk.Button(root, text="Add Model", command=add_model, width=20, height=2)
button_add_model.grid(row=2, column=1, padx=20, pady=10)

# Buttons for updating quantity
button_add = tk.Button(root, text="Add Quantity", command=lambda: update_quantity(True), width=20, height=2)
button_add.grid(row=5, column=0, padx=20, pady=10)

button_subtract = tk.Button(root, text="Subtract Quantity", command=lambda: update_quantity(False), width=20, height=2)
button_subtract.grid(row=5, column=1, padx=20, pady=10)

# Start the GUI event loop
root.mainloop()
