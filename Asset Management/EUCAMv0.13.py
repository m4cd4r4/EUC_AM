# In this code, I have applied the suggested color theme to the background, buttons, 
# and Treeview. The fonts are left as default, but you can change the font in the style 
# configurations for "TButton" if desired. The Combobox, Entry widgets, and Buttons 
# have been given the specified styles for a consistent look. The Treeview selection color 
# is also modified to match the button color.import tkinter as tk
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook

# Load Excel workbook and worksheet
workbook = load_workbook('assets.xlsx')
sheet = workbook.active

# Function to get asset types from the sheet
def get_asset_types():
    return [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]

# Function to refresh the Treeview with the current state of the sheet
def refresh_sheet_view():
    for row in tree.get_children():
        tree.delete(row)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', tk.END, values=row)

# Placeholder functions
def add_new_asset_type():
    # Function implementation
    # After updating sheet
    refresh_sheet_view()

def add_model():
    # Function implementation
    # After updating sheet
    refresh_sheet_view()

def update_quantity(addition):
    # Function implementation
    # After updating sheet
    refresh_sheet_view()

# Initialize Tkinter root
root = tk.Tk()
root.title("EUC Asset Management GUI")
root.geometry("800x600")

# Set a color theme
bg_color = "#f0f0f0"  # Light grey background color
button_color = "#0078D7"  # Blue button color
text_color = "#ffffff"  # White text color

root.configure(bg=bg_color)  # Set the background color of the root window

# Create a style object
style = ttk.Style(root)
style.theme_use("clam")

# Style the Treeview
style.configure("Treeview", background="white", fieldbackground="white", foreground="black")
style.map("Treeview", background=[('selected', button_color)])  # Change selection color

# Style the Buttons
style.configure("TButton", font=('Helvetica', 10), background=button_color, foreground=text_color)
style.map("TButton",
          foreground=[('pressed', text_color), ('active', text_color)],
          background=[('pressed', '!disabled', button_color), ('active', button_color)],
          relief=[('pressed', 'sunken'), ('!pressed', 'raised')])

# Style Combobox
style.configure("TCombobox", fieldbackground=bg_color, background=bg_color)

# Create a frame to hold the widgets
frame = tk.Frame(root, bg=bg_color)
frame.place(relx=0.1, rely=0.1, relwidth=0.8, relheight=0.8)

# Banner label at the top
banner = tk.Label(frame, text="EUC Asset Management GUI", font=("Arial", 16), bg=bg_color, fg='black')
banner.grid(row=0, columnspan=2, pady=20)

# Dropdown for assets
combobox_assets = ttk.Combobox(frame, values=get_asset_types(), width=40, style="TCombobox")
combobox_assets.grid(row=1, columnspan=2, padx=20, pady=10)

# Entry field for new asset type
entry_new_asset = tk.Entry(frame, width=40, bg="white", insertbackground="black")
entry_new_asset.grid(row=2, columnspan=2, padx=20, pady=10)

# Button for adding new asset type
button_add_asset = ttk.Button(frame, text="Add Asset Type", command=add_new_asset_type, style="TButton")
button_add_asset.grid(row=3, column=0, padx=20, pady=10)

# Button for adding model
button_add_model = ttk.Button(frame, text="Add Model", command=add_model, style="TButton")
button_add_model.grid(row=3, column=1, padx=20, pady=10)

# Entry field for quantity
entry_quantity = tk.Entry(frame, width=40, bg="white", insertbackground="black")
entry_quantity.grid(row=4, columnspan=2, padx=20, pady=10)

# Buttons for updating quantity
button_add = ttk.Button(frame, text="Add Quantity", command=lambda: update_quantity(True), style="TButton")
button_add.grid(row=5, column=0, padx=20, pady=10)

button_subtract = ttk.Button(frame, text="Subtract Quantity", command=lambda: update_quantity(False), style="TButton")
button_subtract.grid(row=5, column=1, padx=20, pady=10)

# Treeview for displaying Excel sheet data
columns = [column for column in sheet.iter_rows(max_row=1, values_only=True)][0]
tree = ttk.Treeview(frame, columns=columns, show='headings', style="Treeview")
for col in columns:
    tree.heading(col, text=col)
tree.grid(row=6, columnspan=2, sticky='nsew', pady=(10, 0))

# Initialize the Treeview with sheet data
refresh_sheet_view()

# Start the GUI event loop
root.mainloop()
