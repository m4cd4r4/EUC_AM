import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

# Load Excel workbook and worksheet
workbook = load_workbook('assets.xlsx')
sheet = workbook.active

# History for undo functionality
history = []

# Function to refresh the list of assets
def refresh_assets_list():
    listbox_assets.delete(0, tk.END)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        asset_name, quantity = row[0], row[1]
        listbox_assets.insert(tk.END, f"{asset_name} (Quantity: {quantity})")

# Function to update quantity
def update_quantity(add=True):
    try:
        selected_index = listbox_assets.curselection()[0]
        selected_asset = sheet[f'A{selected_index + 2}'].value
        change = int(entry_quantity.get())
        current_quantity = sheet[f'B{selected_index + 2}'].value

        # Save current state to history for undo
        history.append((selected_index + 2, current_quantity))

        if add:
            sheet[f'B{selected_index + 2}'] = current_quantity + change
        else:
            if current_quantity >= change:
                sheet[f'B{selected_index + 2}'] = current_quantity - change
            else:
                messagebox.showwarning("Warning", "Not enough quantity to remove.")
                return

        workbook.save('assets.xlsx')
        refresh_assets_list()
    except IndexError:
        messagebox.showwarning("Warning", "No asset selected.")
    except ValueError:
        messagebox.showwarning("Warning", "Invalid quantity.")

# Function to undo the last action
def undo_last_action():
    if not history:
        messagebox.showwarning("Warning", "No action to undo.")
        return

    last_action = history.pop()
    row_number, previous_quantity = last_action
    sheet[f'B{row_number}'] = previous_quantity

    workbook.save('assets.xlsx')
    refresh_assets_list()

# Initialize Tkinter root
root = tk.Tk()
root.title("IT Asset Management")

# Listbox to display assets
listbox_assets = tk.Listbox(root, width=50)
listbox_assets.grid(row=0, columnspan=3, padx=10, pady=10)

# Entry field for quantity
entry_quantity = tk.Entry(root)
entry_quantity.grid(row=1, columnspan=3, padx=10)

# Add, Subtract, and Undo buttons
button_add = tk.Button(root, text="+", command=lambda: update_quantity(add=True))
button_add.grid(row=2, column=0, padx=10, pady=10)

button_subtract = tk.Button(root, text="-", command=lambda: update_quantity(add=False))
button_subtract.grid(row=2, column=1, padx=10, pady=10)

button_undo = tk.Button(root, text="Undo", command=undo_last_action)
button_undo.grid(row=2, column=2, padx=10, pady=10)

# Refresh the list initially
refresh_assets_list()

# Start the GUI event loop
root.mainloop()