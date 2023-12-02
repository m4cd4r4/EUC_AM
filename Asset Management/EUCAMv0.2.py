import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

# Load Excel workbook and worksheet
workbook = load_workbook('assets.xlsx')
sheet = workbook.active

# History for undo functionality
history = []

# Function to refresh the list of assets (Updated to include asset labels)
def refresh_assets_list():
    listbox_assets.delete(0, tk.END)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        asset_name, quantity, asset_label = row[0], row[1], row[2] # Now includes asset label
        listbox_assets.insert(tk.END, f"{asset_name} (Quantity: {quantity}, Label: {asset_label})")

# Function to update quantity (Unchanged)
# ...

# Function to update asset label (New Function)
def update_asset_label():
    try:
        selected_index = listbox_assets.curselection()[0]
        new_label = entry_label.get()
        
        # Save current state to history for undo (including label)
        current_label = sheet[f'C{selected_index + 2}'].value
        history.append((selected_index + 2, sheet[f'B{selected_index + 2}'].value, current_label))

        sheet[f'C{selected_index + 2}'] = new_label
        workbook.save('assets.xlsx')
        refresh_assets_list()
    except IndexError:
        messagebox.showwarning("Warning", "No asset selected.")
    except ValueError:
        messagebox.showwarning("Warning", "Invalid label.")

# Function to undo the last action (Updated for labels)
def undo_last_action():
    if not history:
        messagebox.showwarning("Warning", "No action to undo.")
        return

    last_action = history.pop()
    row_number, previous_quantity, previous_label = last_action
    sheet[f'B{row_number}'] = previous_quantity
    sheet[f'C{row_number}'] = previous_label

    workbook.save('assets.xlsx')
    refresh_assets_list()

# Initialize Tkinter root
root = tk.Tk()
root.title("IT Asset Management")

# Listbox, Entry fields, and Buttons setup (Updated for label entry)
# ...

# Entry field for asset label (New Entry Field)
entry_label = tk.Entry(root)
entry_label.grid(row=3, columnspan=3, padx=10)

# Update and Add Label buttons (New Button for updating label)
button_update_label = tk.Button(root, text="Update Label", command=update_asset_label)
button_update_label.grid(row=4, columnspan=3, padx=10, pady=10)

# Refresh the list initially and start the GUI event loop
# ...

(root.mainloop())