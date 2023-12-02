import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

# Load Excel workbook and worksheet
workbook = load_workbook('assets.xlsx')
sheet = workbook.active

# History for undo functionality
history = []

# Function to refresh the list of assets, now accepts listbox_assets as a parameter
def refresh_assets_list(listbox_assets):
    listbox_assets.delete(0, tk.END)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        asset_name, quantity, asset_label = row
        listbox_assets.insert(tk.END, f"{asset_name} (Quantity: {quantity}, Label: {asset_label})")

# Function to update quantity
# Function now needs to accept listbox_assets as a parameter
def update_quantity(listbox_assets):
    try:
        selected_index = listbox_assets.curselection()[0]
        new_quantity = int(entry_quantity.get())

        current_quantity = sheet[f'B{selected_index + 2}'].value
        history.append((selected_index + 2, current_quantity, sheet[f'C{selected_index + 2}'].value))

        sheet[f'B{selected_index + 2}'] = new_quantity
        workbook.save('assets.xlsx')
        refresh_assets_list(listbox_assets)
    except IndexError:
        messagebox.showwarning("Warning", "No asset selected.")
    except ValueError:
        messagebox.showwarning("Warning", "Invalid quantity.")

# Function to update asset label
# ...

# Function to undo the last action
# ...

# Initialize Tkinter root
root = tk.Tk()
root.title("IT Asset Management")

# Listbox to display assets
listbox_assets = tk.Listbox(root, width=50)
listbox_assets.grid(row=0, columnspan=3, padx=10, pady=10)

# Entry field for quantity
entry_quantity = tk.Entry(root)
entry_quantity.grid(row=1, columnspan=3, padx=10)

# Entry field for asset label
entry_label = tk.Entry(root)
entry_label.grid(row=3, columnspan=3, padx=10)

# Buttons for updating quantity and label
# Update the command to pass listbox_assets as an argument
button_update_quantity = tk.Button(root, text="Update Quantity", command=lambda: update_quantity(listbox_assets))
button_update_quantity.grid(row=2, columnspan=3, padx=10, pady=10)

button_update_label = tk.Button(root, text="Update Label", command=lambda: update_asset_label(listbox_assets))
button_update_label.grid(row=4, columnspan=3, padx=10, pady=10)

# Refresh the list initially and start the GUI event loop
# Pass listbox_assets when calling refresh_assets_list
refresh_assets_list(listbox_assets)
root.mainloop()
