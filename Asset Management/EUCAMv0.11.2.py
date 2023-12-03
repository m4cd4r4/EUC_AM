import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook

# Load Excel workbook and worksheet
workbook = load_workbook('assets.xlsx')
sheet = workbook.active

# Function to get asset types from the sheet
def get_asset_types():
    return [row[0].value for row in sheet.iter_rows(min_row=2, values_only=True)]

# Function to refresh the Treeview with the current state of the sheet
def refresh_sheet_view():
    for row in tree.get_children():
        tree.delete(row)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        tree.insert('', tk.END, values=row)

# Placeholder functions
def add_new_asset_type():
    # Function implementation
    # After updating sheet
    refresh_sheet_view()

def add_model():
    # Function implementation
    # After updating sheet
    refresh_sheet_view()

def update_quantity(addition):
    # Function implementation
    # After updating sheet
    refresh_sheet_view()

# Initialize Tkinter root
root = tk.Tk()
root.title("EUC Asset Management GUI")
root.geometry("800x600")

# Create a frame to hold the widgets
frame = tk.Frame(root)
frame.place(relx=0.1, rely=0.1, relwidth=0.8, relheight=0.8)

# ... (rest of your existing widgets setup)

# Treeview for displaying Excel sheet data
columns = [column for column in sheet.iter_rows(max_row=1, values_only=True)][0]
tree = ttk.Treeview(frame, columns=columns, show='headings')
for col in columns:
    tree.heading(col, text=col)
tree.grid(row=6, columnspan=2, sticky='nsew')

# Initialize the Treeview with sheet data
refresh_sheet_view()

# Start the GUI event loop
root.mainloop()
