# Adding Serial # & ServiceNow # to the log_change function, which will write them to the Timestamp sheets, columns E & F

# Build Room\build_roomv2.20.py
# Author: Macdara O Murchu
# 21.02.24

#Reverted again to v2.13 from v2.18 - Loads of broken parts

import logging.config
from pathlib import Path
from tkinter import Menu
import customtkinter as ctk
import os
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook, Workbook
from datetime import datetime
import subprocess
import tkinter.simpledialog as sd

logging_conf_path = Path('logging.conf')
if logging_conf_path.exists() and logging_conf_path.stat().st_size > 0:
    try:
        logging.config.fileConfig(logging_conf_path)
    except Exception as e:
        logging.error(f"Error configuring logging: {e}", exc_info=True)
else:
    logging.basicConfig(level=logging.DEBUG)

def run_inventory_script():
    script_path = script_directory / "inventory-levels_4.2v1.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_4.2v1.py' does not exist in the directory.")

def run_build_room_inventory_script():
    script_path = script_directory / "inventory-levels_BRv1.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_BRv1.py' does not exist in the directory.")

def run_combined_rooms_inventory_script():
    script_path = script_directory / "inventory-levels_combinedv1.1.py"
    if script_path.exists():
        os.system(f"python {script_path}")
    else:
        tk.messagebox.showerror("Error", "The script 'inventory-levels_combinedv1.py' does not exist in the directory.")

def view_headsets_log():
    log_window = tk.Toplevel(root)
    log_window.title("Headsets In Stock")
    log_window.geometry("600x400")

    # Create a Treeview widget to display the log
    columns = ("Serial #", "ServiceNow #")
    log_tree = ttk.Treeview(log_window, columns=columns, show="headings")
    for col in columns:
        log_tree.heading(col, text=col)
        log_tree.column(col, anchor="center")
    log_tree.pack(expand=True, fill="both", padx=10, pady=10)

    # Scrollbar for the Treeview
    scrollbar = ttk.Scrollbar(log_window, orient="vertical", command=log_tree.yview)
    scrollbar.pack(side="right", fill="y")
    log_tree.configure(yscrollcommand=scrollbar.set)

    # Load and display data from the "Headsets" sheet
    if 'Headsets' in workbook.sheetnames:
        headsets_sheet = workbook['Headsets']
        for row in headsets_sheet.iter_rows(min_row=2, values_only=True):
            log_tree.insert('', 'end', values=row)
    else:
        tk.messagebox.showinfo("Info", "Headsets log is empty.", parent=log_window)

def view_all_sans_log():
    log_window = tk.Toplevel(root)
    log_window.title("SANs In Stock")
    log_window.geometry("600x400")

    # Create a Treeview widget to display the log
    columns = ("SAN Number", "Item", "Timestamp")
    log_tree = ttk.Treeview(log_window, columns=columns, show="headings")
    for col in columns:
        log_tree.heading(col, text=col)
        log_tree.column(col, anchor="center")
    log_tree.pack(expand=True, fill="both", padx=10, pady=10)

    # Scrollbar for the Treeview
    scrollbar = ttk.Scrollbar(log_window, orient="vertical", command=log_tree.yview)
    scrollbar.pack(side="right", fill="y")
    log_tree.configure(yscrollcommand=scrollbar.set)

    # Load and display data from the "All SANs" sheet
    if 'All SANs' in workbook.sheetnames:
        all_sans_sheet = workbook['All SANs']
        for row in all_sans_sheet.iter_rows(min_row=2, values_only=True):
            log_tree.insert('', 'end', values=row)
    else:
        tk.messagebox.showinfo("Info", "All SANs log is empty.", parent=log_window)

root = ctk.CTk()
root.title("Perth EUC Assets")
root.geometry("650x650")

menu_bar = tk.Menu(root)
plots_menu = tk.Menu(menu_bar, tearoff=0)
plots_menu.add_command(label="Basement 4.2 Inventory", command=run_inventory_script)
plots_menu.add_command(label="Build Room Inventory", command=run_build_room_inventory_script)
plots_menu.add_command(label="Combined Inventory", command=run_combined_rooms_inventory_script)
plots_menu.add_command(label="SANs In Stock", command=view_all_sans_log)
plots_menu.add_command(label="Headsets In Stock", command=view_headsets_log)
menu_bar.add_cascade(label="Data", menu=plots_menu)
root.config(menu=menu_bar)

script_directory = Path(__file__).parent
workbook_path = script_directory / 'EUC_Perth_Assets.xlsx'
if Path(workbook_path).exists():
    workbook = load_workbook(workbook_path)
else:
    workbook = Workbook()
    workbook.active.title = '4.2 Items'
    workbook.create_sheet('4.2 Timestamps')
    workbook.create_sheet('BR Items')
    workbook.create_sheet('BR Timestamps')
    workbook.create_sheet('Project Designated Items')
    workbook.create_sheet('Project Designated Timestamps')
    workbook.create_sheet('All SANs')
    workbook['4.2 Items'].append(["Item", "LastCount", "NewCount"])
    workbook['4.2 Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['BR Items'].append(["Item", "LastCount", "NewCount"])
    workbook['BR Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['Project Designated Items'].append(["Item", "LastCount", "NewCount"])
    workbook['Project Designated Timestamps'].append(["Timestamp", "Item", "Action", "SAN Number"])
    workbook['All SANs'].append(["SAN Number", "Item", "Timestamp"])
    workbook.save(workbook_path)

all_sans_sheet = workbook['All SANs']
sheets = {'original': ('4.2 Items', '4.2 Timestamps'), 'backup': ('BR Items', 'BR Timestamps')}
current_sheets = sheets['original']

style = ttk.Style()
style.configure("Treeview", font=('Helvetica', 12))

vcmd = (root.register(lambda P: P.isdigit() or P == ""), '%P')

class SANInputDialog(tk.Toplevel):
    def __init__(self, parent, title=None):
        super().__init__(parent)
        self.transient(parent)
        self.title(title)
        self.parent = parent
        self.result = None
        self.create_widgets()
        self.grab_set()
        self.geometry(f"+{parent.winfo_rootx() + parent.winfo_width() // 2 - 100}+{parent.winfo_rooty() + parent.winfo_height() // 2 - 50}")
        self.wait_window(self)

    def create_widgets(self):
        self.entry = ttk.Entry(self, validate="key", validatecommand=vcmd)
        self.entry.pack(padx=5, pady=5)
        button_frame = tk.Frame(self)
        button_frame.pack(pady=5)
        submit_button = ttk.Button(button_frame, text="Submit", command=self.on_submit)
        submit_button.pack(side='left', padx=5)
        cancel_button = ttk.Button(button_frame, text="Cancel", command=self.on_cancel)
        cancel_button.pack(side='left', padx=5)

    def on_submit(self):
        san_input = self.entry.get()
        if san_input and len(san_input) >= 5 and len(san_input) <= 6:
            self.result = san_input
            self.destroy()
        else:
            tk.messagebox.showerror("Error", "Please enter a valid SAN number.", parent=self)
            self.entry.focus_set()

    def on_cancel(self):
        self.result = None
        self.destroy()

def is_san_unique(san_number):
    # Adjust the search to account for the 'SAN' prefix properly
    search_string = "SAN" + san_number if not san_number.startswith("SAN") else san_number
    unique = all(search_string != row[0] for row in all_sans_sheet.iter_rows(min_row=2, values_only=True))
    print(f"Checking SAN {search_string}: Unique - {unique}")  # Debug print
    return unique

def show_san_input():
    dialog = SANInputDialog(root, "SAN #")
    return dialog.result

def open_spreadsheet():
    try:
        if os.name == 'nt':
            os.startfile(workbook_path)
        else:
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.run([opener, workbook_path])
    except Exception as e:
        tk.messagebox.showerror("Error", f"Failed to open the spreadsheet: {e}")

frame = ctk.CTkFrame(root)
frame.pack(padx=3, pady=3, fill='both', expand=True)
entry_frame = ctk.CTkFrame(frame)
entry_frame.pack(pady=3)

plots_menu.add_command(label="Open Spreadsheet", command=open_spreadsheet)

button_width = 25
button_1 = ctk.CTkButton(entry_frame, text="Basement 4.2", command=lambda: switch_sheets('original'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_1.pack(side='left', padx=3)
button_2 = ctk.CTkButton(entry_frame, text="Build Room", command=lambda: switch_sheets('backup'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_2.pack(side='left', padx=(3, 50))
button_subtract = ctk.CTkButton(entry_frame, text="-", command=lambda: update_count('subtract'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_subtract.pack(side='left', padx=3)
entry_value = tk.Entry(entry_frame, font=("Helvetica", 14), justify='center', width=5, validate="key", validatecommand=vcmd)
entry_value.pack(side='left', padx=3)
button_add = ctk.CTkButton(entry_frame, text="+", command=lambda: update_count('add'), width=button_width, font=("Helvetica", 14), corner_radius=3)
button_add.pack(side='left', padx=3)

def update_treeview():
    tree.delete(*tree.get_children())
    workbook = load_workbook(workbook_path)
    item_sheet = workbook[current_sheets[0]]
    row_count = 0
    for row in item_sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            if row_count % 2 == 0:
                bg_color = 'white'
            else:
                bg_color = '#f0f0f0'
            tree.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
            tree.tag_configure('oddrow', background='#f0f0f0')
            tree.tag_configure('evenrow', background='white')
            row_count += 1

def log_change(item, action, count=1, san_number="", timestamp_sheet=None, serial_num="", servicenow_num=""):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    action_text = f"{action} {count}" if action in ['add', 'subtract'] else action
    try:
        if timestamp_sheet is not None:
            # Updated to include serial_num and servicenow_num in the row.
            # Append a new row with the relevant information to the specified timestamp sheet.
            timestamp_sheet.append([timestamp, item, action_text, san_number, serial_num, servicenow_num])
            workbook.save(workbook_path)
            # Ensure the log_view is updated to reflect the recent changes.
            update_log_view()
            # Log the operation for debugging or auditing purposes.
            logging.info(f"Logged change: Time: {timestamp}, Item: {item}, Action: {action_text}, SAN: {san_number}, Serial#: {serial_num}, ServiceNow#: {servicenow_num}")
        else:
            logging.error("No timestamp sheet provided for logging.")
    except Exception as e:
        logging.error(f"Failed to log change: {e}")
        tk.messagebox.showerror("Error", f"Failed to log change: {e}")

def switch_sheets(sheet_type):
    global current_sheets
    current_sheets = sheets[sheet_type]
    update_treeview()
    update_log_view()

def update_log_view():
    if 'log_view' in globals():
        log_view.delete(*log_view.get_children())
        log_sheet = workbook[current_sheets[1]]
        all_rows = list(log_sheet.iter_rows(min_row=2, values_only=True))
        # Adjust the sorting to use the first column (timestamp)
        sorted_rows = sorted(all_rows, key=lambda r: datetime.strptime(r[0], "%Y-%m-%d %H:%M:%S") if r[0] else datetime.min, reverse=True)
        row_count = 0
        for row in sorted_rows:
            if row[0] is not None:
                log_view.insert('', 'end', values=row, tags=('oddrow' if row_count % 2 == 1 else 'evenrow'))
                log_view.tag_configure('oddrow', background='#f0f0f0')
                log_view.tag_configure('evenrow', background='white')
                row_count += 1

# Serial Number Input Dialog Function
def serial_number_input():
    while True:
        serial_num = sd.askstring("Serial Number", "Enter Serial Number:", parent=root)
        if serial_num is None:  # User pressed cancel
            return None
        if len(serial_num) == 6 and serial_num.isalnum():
            return serial_num  # Valid input
        else:
            tk.messagebox.showerror("Invalid Input", "Serial Number must be 6 characters.")

class ServiceNowInputDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.result = None
        self.title("ServiceNow #")

        self.init_ui()

    def init_ui(self):
        # Frame for input fields
        input_frame = tk.Frame(self)
        input_frame.pack(padx=10, pady=10, fill=tk.X, expand=True)

        # Prefix selection
        tk.Label(input_frame, text="Prefix:").grid(row=0, column=0, sticky="w")
        self.prefix_var = tk.StringVar(self)
        self.prefix_combobox = ttk.Combobox(input_frame, textvariable=self.prefix_var, state="readonly", values=["TASK", "RITM"])
        self.prefix_combobox.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        self.prefix_combobox.set("TASK")  # Default value

        # ServiceNow number entry
        tk.Label(input_frame, text="Number:").grid(row=1, column=0, sticky="w")
        self.service_now_entry = ttk.Entry(input_frame)
        self.service_now_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=5)

        # Button Frame
        button_frame = tk.Frame(self)
        button_frame.pack(fill=tk.X, expand=True, padx=10, pady=10)
        submit_btn = ttk.Button(button_frame, text="Submit", command=self.submit)
        submit_btn.pack(side=tk.RIGHT, padx=5)
        cancel_btn = ttk.Button(button_frame, text="Cancel", command=self.cancel)
        cancel_btn.pack(side=tk.RIGHT, padx=5)

    def submit(self):
        prefix = self.prefix_var.get()
        number = self.service_now_entry.get()
        if number.isdigit() and 6 <= len(number) <= 8:
            self.result = f"{prefix}{number}"
            self.destroy()
        else:
            tk.messagebox.showerror("Error", "Please enter a valid ServiceNow number.", parent=self)

    def cancel(self):
        self.result = None
        self.destroy()

    def show_dialog(self):
        self.wait_window(self)
        return self.result
    
def servicenow_number_input():
    dialog = ServiceNowInputDialog(root)
    return dialog.show()

def update_count(operation):
    selected_item = tree.item(tree.focus())['values'][0] if tree.focus() else None
    if selected_item and "Headset" in selected_item:
        serial_number = serial_number_input()
        if serial_number:
            servicenow_number = servicenow_number_input()
            if servicenow_number:  # Check if user provided a ServiceNow number
                # Proceed with original functionality as the user did not cancel the dialog
                headsets_sheet = workbook['Headsets']  # Assuming 'headsets' sheet exists
                headsets_sheet.append([serial_number, servicenow_number])  # Now appending both serial and ServiceNow numbers together
                workbook.save(workbook_path)
            else:
                # Here we do not proceed with any changes since the servicenow dialog was cancelled
                tk.messagebox.showinfo("Operation Cancelled", "The operation was cancelled. No changes have been made.")
                return  # Stop the function execution as the servicenow dialog was cancelled

            workbook.save(workbook_path)
    
    if selected_item:
        input_value = entry_value.get()
        if input_value.isdigit():
            input_value = int(input_value)
            item_sheet = workbook[current_sheets[0]]
            timestamp_sheet = workbook[current_sheets[1]]
            san_required = any(g in selected_item for g in ["G8", "G9", "G10"])

            if san_required:
                san_count = 0
                while san_count < input_value:
                    san_number = show_san_input()
                    if san_number is None:  # User cancelled the input
                        return
                    san_number = "SAN" + san_number if not san_number.startswith("SAN") else san_number

                    if operation == 'add':
                        if is_san_unique(san_number):
                            print(f"Adding unique SAN {san_number}")  # Debug print
                            all_sans_sheet.append([san_number, selected_item, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                            log_change(selected_item, operation, 1, san_number, timestamp_sheet)  # Correctly pass the SAN number
                            san_count += 1
                        else:
                            tk.messagebox.showerror("Error", f"Duplicate or already used SAN number: {san_number}", parent=root)
                    elif operation == 'subtract':
                        # Check if the SAN number exists in the sheet
                        row_to_delete = None
                        for row in all_sans_sheet.iter_rows(min_row=2):
                            if row[0].value == san_number:
                                row_to_delete = row[0].row
                                break

                        if row_to_delete:
                            all_sans_sheet.delete_rows(row_to_delete)
                            log_change(selected_item, operation, 1, san_number, timestamp_sheet)
                            san_count += 1
                        else:
                            tk.messagebox.showerror("Error", f"The following SAN number is not in the inventory: {san_number}", parent=root)

            # Adjust item counts
            for row in item_sheet.iter_rows(min_row=2):
                if row[0].value == selected_item:
                    row[1].value = row[2].value or 0
                    if operation == 'add':
                        row[2].value = (row[2].value or 0) + input_value
                    elif operation == 'subtract':
                        row[2].value = max((row[2].value or 0) - input_value, 0)

            # Log the change for items not requiring SAN
            if not san_required:
                log_change(selected_item, operation, input_value, "", timestamp_sheet)

            workbook.save(workbook_path)
            update_treeview()
            update_log_view()

columns = ("Item", "LastCount", "NewCount")
tree = ttk.Treeview(frame, columns=columns, show="headings", selectmode='browse', style="Treeview")
for col in columns:
    tree.heading(col, text=col, anchor='center')
    tree.column("Item", anchor='center', width=250, stretch=False) # Width of the "Item" column in the treeview. The other columns are default width.
    tree.column("LastCount", anchor='center', width=175, stretch=False)
    tree.column("NewCount", anchor='center', width=175, stretch=False)
tree.pack(expand=True, fill="both", padx=3, pady=3)

log_view_frame = ctk.CTkFrame(root)
log_view_frame.pack(side=tk.BOTTOM, fill='both', expand=True, padx=10, pady=10)

log_view_columns = ("Timestamp", "Item", "Action", "SAN Number", "Serial #", "ServiceNow #")
log_view = ttk.Treeview(log_view_frame, columns=log_view_columns, show="headings", style="Treeview", height=8)
for col in log_view_columns:
    log_view.heading(col, text=col, anchor='center')
    log_view.column("Timestamp", anchor='center', width=190, stretch=False)
    log_view.column("Item", anchor='center', width=160, stretch=False)
    log_view.column("Action", anchor='center', width=100, stretch=False)
    log_view.column("SAN Number", anchor='center', width=100, stretch=False)
    log_view.column("Serial #", anchor='center', width=100, stretch=False)
    log_view.column("ServiceNow #", anchor='center', width=100, stretch=False)

scrollbar_log = ttk.Scrollbar(log_view_frame, orient="vertical", command=log_view.yview)
scrollbar_log.pack(side='right', fill='y')
log_view.configure(yscrollcommand=scrollbar_log.set)
log_view.pack(expand=True, fill='both')

root.after(100, update_treeview)
update_log_view()

root.mainloop()